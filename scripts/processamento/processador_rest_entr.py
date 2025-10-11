#!/usr/bin/env python3
"""
PROCESSADOR REST_ENTR(DIA) - Próxima tabela do caixa
Explora e processa tabela rest_entr(dia) para complementar os dados de vendas
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime

class ProcessadorRestEntr:
    def __init__(self):
        self.pasta_dados = Path("data/caixa_lojas")
        self.pasta_resultados = Path("data/rest_entr_resultados")
        self.pasta_resultados.mkdir(parents=True, exist_ok=True)
        
        self.lojas_ativas = ['MAUA', 'SUZANO', 'RIO_PEQUENO', 'PERUS', 'SUZANO2', 'SAO_MATEUS']
    
    def investigar_rest_entr_arquivo(self, arquivo_path):
        """Investiga onde está a tabela rest_entr(dia) no arquivo"""
        print(f"   🔍 Investigando: {arquivo_path.name}")
        
        try:
            wb = openpyxl.load_workbook(arquivo_path, data_only=True)
            
            tabelas_encontradas = []
            
            # Buscar por "rest" e "entr" em todas as sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Verificar se é sheet específica para rest_entr
                if 'rest' in sheet_name.lower() and 'entr' in sheet_name.lower():
                    tabelas_encontradas.append(f"Sheet específica: {sheet_name}")
                
                # Buscar por células que contenham "rest" e "entr"
                for row in range(1, min(50, ws.max_row + 1)):
                    for col in range(1, min(20, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_lower = cell_value.lower()
                            if ('rest' in cell_lower and 'entr' in cell_lower) or \
                               'restante entrada' in cell_lower or \
                               'rest_entr' in cell_lower:
                                coord = ws.cell(row=row, column=col).coordinate
                                tabelas_encontradas.append(f"Sheet {sheet_name}, {coord}: {cell_value}")
            
            wb.close()
            
            if tabelas_encontradas:
                print(f"      ✅ Encontradas {len(tabelas_encontradas)} referências:")
                for tabela in tabelas_encontradas[:3]:  # Mostrar só as primeiras 3
                    print(f"         📍 {tabela}")
            else:
                print(f"      ❌ Nenhuma referência rest_entr encontrada")
            
            return tabelas_encontradas
            
        except Exception as e:
            print(f"      ❌ Erro: {e}")
            return []
    
    def investigar_estrutura_rest_entr_loja(self, loja):
        """Investiga estrutura rest_entr em uma loja específica"""
        print(f"\n🔍 INVESTIGANDO REST_ENTR: {loja}")
        print("=" * 60)
        
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            print(f"❌ Loja {loja} não encontrada")
            return
        
        arquivos_investigados = 0
        total_referencias = 0
        
        # Investigar alguns arquivos da loja
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~') and arquivos_investigados < 3:
                referencias = self.investigar_rest_entr_arquivo(arquivo)
                total_referencias += len(referencias)
                arquivos_investigados += 1
        
        # Investigar pastas de anos
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir() and any(ano in pasta.name for ano in ['2023', '2024', '2025']) and arquivos_investigados < 5:
                for arquivo in pasta.glob("*.xlsx"):
                    if not arquivo.name.startswith('~') and arquivos_investigados < 5:
                        referencias = self.investigar_rest_entr_arquivo(arquivo)
                        total_referencias += len(referencias)
                        arquivos_investigados += 1
                        break
        
        print(f"\n📊 RESUMO {loja}:")
        print(f"   Arquivos investigados: {arquivos_investigados}")
        print(f"   Total referências rest_entr: {total_referencias}")
        
        return total_referencias
    
    def extrair_rest_entr_detalhado(self, arquivo_path, loja):
        """Extrai dados detalhados da tabela rest_entr quando encontrada"""
        print(f"   📄 Processando rest_entr: {arquivo_path.name}")
        
        try:
            wb = openpyxl.load_workbook(arquivo_path, data_only=True)
            
            # Detectar ano e mês
            nome_arquivo = arquivo_path.name.lower()
            meses = {'jan': 'JAN', 'fev': 'FEV', 'mar': 'MAR', 'abr': 'ABR', 
                    'mai': 'MAI', 'jun': 'JUN', 'jul': 'JUL', 'ago': 'AGO',
                    'set': 'SET', 'out': 'OUT', 'nov': 'NOV', 'dez': 'DEZ'}
            
            mes_nome = 'UNKNOWN'
            for mes_abr, mes_nome_full in meses.items():
                if mes_abr in nome_arquivo:
                    mes_nome = mes_nome_full
                    break
            
            if '_25' in nome_arquivo or '2025' in nome_arquivo:
                ano = '2025'
            elif '_24' in nome_arquivo or '2024' in nome_arquivo:
                ano = '2024'
            elif '_23' in nome_arquivo or '2023' in nome_arquivo:
                ano = '2023'
            else:
                ano = '2024'
            
            dados_rest_entr = []
            
            # Processar cada dia (sheets numeradas)
            for dia in range(1, 32):
                dia_str = f"{dia:02d}"
                
                if dia_str in wb.sheetnames:
                    ws = wb[dia_str]
                    
                    # Buscar por região de rest_entr
                    for row in range(1, min(50, ws.max_row + 1)):
                        for col in range(1, min(20, ws.max_column + 1)):
                            cell_value = ws.cell(row=row, column=col).value
                            
                            if cell_value and isinstance(cell_value, str):
                                cell_lower = cell_value.lower()
                                if ('rest' in cell_lower and 'entr' in cell_lower) or \
                                   'restante entrada' in cell_lower:
                                    
                                    print(f"         📍 Rest_entr encontrado em {dia_str}, linha {row}, coluna {col}: {cell_value}")
                                    
                                    # Tentar extrair dados próximos
                                    inicio_dados = row + 1
                                    entradas_dia = 0
                                    
                                    for data_row in range(inicio_dados, min(inicio_dados + 20, ws.max_row + 1)):
                                        linha_dados = []
                                        tem_dados = False
                                        
                                        for data_col in range(col, min(col + 5, ws.max_column + 1)):
                                            valor = ws.cell(row=data_row, column=data_col).value
                                            linha_dados.append(valor)
                                            if valor and str(valor).strip():
                                                tem_dados = True
                                        
                                        if tem_dados:
                                            dados_rest_entr.append({
                                                'loja': loja,
                                                'ano': ano,
                                                'mes': mes_nome,
                                                'dia': int(dia),
                                                'arquivo': arquivo_path.name,
                                                'linha_origem': data_row,
                                                'dados_brutos': linha_dados,
                                                'cabecalho_encontrado': cell_value
                                            })
                                            entradas_dia += 1
                                    
                                    if entradas_dia > 0:
                                        print(f"            ✅ {entradas_dia} entradas rest_entr extraídas")
                                    
                                    break  # Sair do loop de colunas
                        else:
                            continue
                        break  # Sair do loop de linhas se encontrou
            
            wb.close()
            
            print(f"      📊 TOTAL rest_entr: {len(dados_rest_entr)} registros")
            return dados_rest_entr
            
        except Exception as e:
            print(f"      ❌ Erro: {e}")
            return []
    
    def investigar_todas_lojas_rest_entr(self):
        """Investiga rest_entr em todas as lojas"""
        print("🔍 INVESTIGAÇÃO REST_ENTR - TODAS AS LOJAS")
        print("=" * 60)
        print("🎯 Objetivo: Encontrar e mapear tabela rest_entr(dia)")
        print()
        
        resultados = {}
        
        for loja in self.lojas_ativas:
            referencias = self.investigar_estrutura_rest_entr_loja(loja)
            resultados[loja] = referencias
        
        print(f"\n📊 RESUMO GERAL REST_ENTR:")
        print("=" * 40)
        
        total_referencias = sum(resultados.values())
        lojas_com_rest_entr = len([l for l, r in resultados.items() if r > 0])
        
        print(f"📈 Total referências encontradas: {total_referencias}")
        print(f"🏢 Lojas com rest_entr: {lojas_com_rest_entr}/{len(self.lojas_ativas)}")
        
        print(f"\n📋 DETALHES POR LOJA:")
        for loja, refs in resultados.items():
            status = "✅" if refs > 0 else "❌"
            print(f"   {status} {loja}: {refs} referências")
        
        return resultados

def main():
    processador = ProcessadorRestEntr()
    
    print("🔍 PROCESSADOR REST_ENTR(DIA)")
    print("=" * 50)
    print("🎯 Próxima tabela após vend(dia) processada")
    print("💡 Objetivo: Mapear restante entrada por dia")
    print()
    print("1. Investigar todas as lojas (rest_entr)")
    print("2. Investigar loja específica")
    print("3. Processar dados rest_entr (quando encontrados)")
    print("4. Sair")
    
    while True:
        escolha = input("\n👉 Escolha uma opção (1-4): ").strip()
        
        if escolha == "1":
            processador.investigar_todas_lojas_rest_entr()
            break
        elif escolha == "2":
            loja = input("Digite o nome da loja: ").strip().upper()
            if loja in processador.lojas_ativas:
                processador.investigar_estrutura_rest_entr_loja(loja)
            else:
                print(f"❌ Loja {loja} não encontrada")
            break
        elif escolha == "3":
            print("🚧 Função de processamento será implementada após mapear estrutura")
            break
        elif escolha == "4":
            print("👋 Saindo...")
            break
        else:
            print("❌ Opção inválida")

if __name__ == "__main__":
    main()