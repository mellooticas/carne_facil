#!/usr/bin/env python3
"""
PROCESSADOR FOCO TABELA VEND(DIA) - VERSÃO EXCLUSIVA
Sistema focado APENAS na tabela 'vend(dia)' presente em todos os arquivos Excel
Objetivo: Capturar os R$ 5,7M reais sem inconsistências de outras fontes
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import re

class ProcessadorVendDia:
    def __init__(self):
        self.pasta_dados = Path("data/caixa_lojas")
        self.pasta_resultados = Path("data/vendas_vend_dia")
        self.pasta_resultados.mkdir(parents=True, exist_ok=True)
        
        self.lojas_ativas = ['MAUA', 'SUZANO', 'RIO_PEQUENO', 'PERUS', 'SUZANO2', 'SAO_MATEUS']
        
    def verificar_tabelas_excel(self, arquivo_path):
        """Verifica se existe a tabela vend(dia) no arquivo Excel"""
        try:
            wb = openpyxl.load_workbook(arquivo_path, data_only=True)
            
            # Buscar tabela vend(dia) em todas as sheets
            tabelas_encontradas = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Verificar se existe a tabela vend(dia)
                if 'vend' in sheet_name.lower() and 'dia' in sheet_name.lower():
                    tabelas_encontradas.append(f"Sheet: {sheet_name}")
                
                # Buscar por células que contenham "vend" e "dia"
                for row in range(1, min(20, ws.max_row + 1)):
                    for col in range(1, min(20, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            if 'vend' in cell_value.lower() and 'dia' in cell_value.lower():
                                tabelas_encontradas.append(f"Célula {ws.cell(row=row, column=col).coordinate}: {cell_value}")
            
            wb.close()
            return tabelas_encontradas
            
        except Exception as e:
            return [f"Erro: {e}"]
    
    def extrair_vend_dia_puro(self, arquivo_path, loja):
        """Extrai dados APENAS da tabela vend(dia)"""
        print(f"   📄 Analisando: {arquivo_path.name}")
        
        try:
            wb = openpyxl.load_workbook(arquivo_path, data_only=True)
            
            # Detectar ano e mês do arquivo
            nome_arquivo = arquivo_path.name.lower()
            meses = {'jan': 'JAN', 'fev': 'FEV', 'mar': 'MAR', 'abr': 'ABR', 
                    'mai': 'MAI', 'jun': 'JUN', 'jul': 'JUL', 'ago': 'AGO',
                    'set': 'SET', 'out': 'OUT', 'nov': 'NOV', 'dez': 'DEZ'}
            
            mes_nome = 'UNKNOWN'
            for mes_abr, mes_nome_full in meses.items():
                if mes_abr in nome_arquivo:
                    mes_nome = mes_nome_full
                    break
            
            if '_25' in nome_arquivo or '2025' in nome_arquivo:
                ano = '2025'
            elif '_24' in nome_arquivo or '2024' in nome_arquivo:
                ano = '2024'
            elif '_23' in nome_arquivo or '2023' in nome_arquivo:
                ano = '2023'
            else:
                ano = '2024'
            
            vendas_vend_dia = []
            
            # Verificar se existe sheet específica para vend(dia)
            sheet_vend_dia = None
            for sheet_name in wb.sheetnames:
                if 'vend' in sheet_name.lower() and 'dia' in sheet_name.lower():
                    sheet_vend_dia = sheet_name
                    break
            
            if sheet_vend_dia:
                print(f"      ✅ Encontrada sheet vend(dia): {sheet_vend_dia}")
                ws = wb[sheet_vend_dia]
                
                # Buscar dados na sheet vend(dia)
                for row in range(1, ws.max_row + 1):
                    linha_dados = []
                    for col in range(1, min(15, ws.max_column + 1)):
                        valor = ws.cell(row=row, column=col).value
                        linha_dados.append(valor)
                    
                    # Se a linha tem dados úteis
                    if any(str(v).strip() for v in linha_dados if v):
                        vendas_vend_dia.append({
                            'loja': loja,
                            'arquivo': arquivo_path.name,
                            'sheet': sheet_vend_dia,
                            'linha': row,
                            'dados': linha_dados
                        })
            
            else:
                # Buscar por tabela vend(dia) em sheets de dias
                print(f"      🔍 Buscando tabela vend(dia) em sheets de dias...")
                
                for dia in range(1, 32):
                    dia_str = f"{dia:02d}"
                    
                    if dia_str in wb.sheetnames:
                        ws = wb[dia_str]
                        
                        # Buscar por cabeçalho "vend(dia)" ou similar
                        for row in range(1, min(50, ws.max_row + 1)):
                            for col in range(1, min(20, ws.max_column + 1)):
                                cell_value = ws.cell(row=row, column=col).value
                                
                                if cell_value and isinstance(cell_value, str):
                                    if ('vend' in cell_value.lower() and 'dia' in cell_value.lower()) or \
                                       'vendas do dia' in cell_value.lower() or \
                                       'total do dia' in cell_value.lower():
                                        
                                        print(f"         📍 Tabela encontrada em {dia_str}, linha {row}, coluna {col}: {cell_value}")
                                        
                                        # Extrair dados da tabela vend(dia)
                                        inicio_dados = row + 1
                                        for data_row in range(inicio_dados, min(inicio_dados + 50, ws.max_row + 1)):
                                            linha_dados = []
                                            tem_dados = False
                                            
                                            for data_col in range(col, min(col + 10, ws.max_column + 1)):
                                                valor = ws.cell(row=data_row, column=data_col).value
                                                linha_dados.append(valor)
                                                if valor and str(valor).strip():
                                                    tem_dados = True
                                            
                                            if tem_dados:
                                                vendas_vend_dia.append({
                                                    'loja': loja,
                                                    'arquivo': arquivo_path.name,
                                                    'dia': dia_str,
                                                    'sheet': dia_str,
                                                    'linha_tabela': data_row,
                                                    'dados': linha_dados,
                                                    'cabecalho_encontrado': cell_value
                                                })
            
            wb.close()
            
            print(f"      📊 {len(vendas_vend_dia)} linhas de vend(dia) extraídas")
            return vendas_vend_dia
            
        except Exception as e:
            print(f"      ❌ Erro: {e}")
            return []
    
    def verificar_estrutura_loja(self, loja):
        """Verifica estrutura das tabelas vend(dia) em uma loja"""
        print(f"\n🔍 VERIFICANDO ESTRUTURA VEND(DIA): {loja}")
        print("=" * 60)
        
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            print(f"❌ Loja {loja} não encontrada")
            return
        
        arquivos_verificados = 0
        estruturas_encontradas = []
        
        # Verificar alguns arquivos da loja
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~') and arquivos_verificados < 3:
                print(f"\n📄 Verificando: {arquivo.name}")
                
                tabelas = self.verificar_tabelas_excel(arquivo)
                
                if tabelas:
                    print(f"   ✅ Tabelas encontradas:")
                    for tabela in tabelas:
                        print(f"      📌 {tabela}")
                        estruturas_encontradas.append(f"{arquivo.name}: {tabela}")
                else:
                    print(f"   ❌ Nenhuma tabela vend(dia) encontrada")
                
                arquivos_verificados += 1
        
        # Verificar pastas de anos
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir() and any(ano in pasta.name for ano in ['2023', '2024', '2025']) and arquivos_verificados < 5:
                for arquivo in pasta.glob("*.xlsx"):
                    if not arquivo.name.startswith('~') and arquivos_verificados < 5:
                        print(f"\n📄 Verificando: {pasta.name}/{arquivo.name}")
                        
                        tabelas = self.verificar_tabelas_excel(arquivo)
                        
                        if tabelas:
                            print(f"   ✅ Tabelas encontradas:")
                            for tabela in tabelas:
                                print(f"      📌 {tabela}")
                                estruturas_encontradas.append(f"{arquivo.name}: {tabela}")
                        else:
                            print(f"   ❌ Nenhuma tabela vend(dia) encontrada")
                        
                        arquivos_verificados += 1
                        break
        
        print(f"\n📊 RESUMO ESTRUTURAS {loja}:")
        print(f"   Arquivos verificados: {arquivos_verificados}")
        print(f"   Estruturas encontradas: {len(estruturas_encontradas)}")
        
        return estruturas_encontradas
    
    def processar_loja_vend_dia(self, loja):
        """Processa loja focando APENAS na tabela vend(dia)"""
        print(f"\n📊 PROCESSANDO VEND(DIA): {loja}")
        print("=" * 60)
        
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            print(f"❌ Loja {loja} não encontrada")
            return None
        
        todas_vendas_vend_dia = []
        arquivos_processados = 0
        
        # Processar arquivos da raiz (2025)
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~'):
                vendas = self.extrair_vend_dia_puro(arquivo, loja)
                todas_vendas_vend_dia.extend(vendas)
                if vendas:
                    arquivos_processados += 1
        
        # Processar pastas de anos
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir() and any(ano in pasta.name for ano in ['2023', '2024', '2025']):
                for arquivo in pasta.glob("*.xlsx"):
                    if not arquivo.name.startswith('~'):
                        vendas = self.extrair_vend_dia_puro(arquivo, loja)
                        todas_vendas_vend_dia.extend(vendas)
                        if vendas:
                            arquivos_processados += 1
        
        if todas_vendas_vend_dia:
            # Salvar dados brutos da tabela vend(dia)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"VEND_DIA_RAW_{loja}_{timestamp}.json"
            caminho_arquivo = self.pasta_resultados / nome_arquivo
            
            import json
            with open(caminho_arquivo, 'w', encoding='utf-8') as f:
                json.dump(todas_vendas_vend_dia, f, ensure_ascii=False, indent=2, default=str)
            
            print(f"\n📊 RESULTADO VEND(DIA) {loja}:")
            print(f"   ✅ Arquivos processados: {arquivos_processados}")
            print(f"   📈 Linhas vend(dia): {len(todas_vendas_vend_dia):,}")
            print(f"   💾 Arquivo salvo: {nome_arquivo}")
            
            return {
                'loja': loja,
                'arquivo': nome_arquivo,
                'linhas_vend_dia': len(todas_vendas_vend_dia),
                'arquivos_processados': arquivos_processados
            }
        
        return None

def main():
    processador = ProcessadorVendDia()
    
    print("📊 PROCESSADOR FOCO TABELA VEND(DIA)")
    print("=" * 50)
    print("🎯 Foco EXCLUSIVO na tabela 'vend(dia)' para R$ 5,7M")
    print()
    print("1. Verificar estrutura vend(dia) de uma loja")
    print("2. Processar todas as lojas (vend(dia) apenas)")
    print("3. Processar loja específica")
    print("4. Sair")
    
    while True:
        escolha = input("\n👉 Escolha uma opção (1-4): ").strip()
        
        if escolha == "1":
            loja = input("Digite o nome da loja para verificar: ").strip().upper()
            if loja in processador.lojas_ativas:
                processador.verificar_estrutura_loja(loja)
            else:
                print(f"❌ Loja {loja} não encontrada")
            break
            
        elif escolha == "2":
            print("\n🚀 PROCESSANDO TODAS AS LOJAS - FOCO VEND(DIA)")
            resultados = []
            
            for loja in processador.lojas_ativas:
                resultado = processador.processar_loja_vend_dia(loja)
                if resultado:
                    resultados.append(resultado)
            
            if resultados:
                print(f"\n🎉 PROCESSAMENTO VEND(DIA) CONCLUÍDO!")
                print("=" * 50)
                
                total_linhas = sum(r['linhas_vend_dia'] for r in resultados)
                
                print(f"📊 RESUMO VEND(DIA):")
                print(f"   🏢 Lojas processadas: {len(resultados)}")
                print(f"   📈 Total linhas vend(dia): {total_linhas:,}")
                
                for resultado in resultados:
                    print(f"   ✅ {resultado['loja']}: {resultado['linhas_vend_dia']} linhas")
            break
            
        elif escolha == "3":
            loja = input("Digite o nome da loja: ").strip().upper()
            if loja in processador.lojas_ativas:
                processador.processar_loja_vend_dia(loja)
            else:
                print(f"❌ Loja {loja} não encontrada")
            break
            
        elif escolha == "4":
            print("👋 Saindo...")
            break
        else:
            print("❌ Opção inválida")

if __name__ == "__main__":
    main()