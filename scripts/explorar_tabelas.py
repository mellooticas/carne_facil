"""
Script para explorar tabelas Excel e sua estrutura
Específico para identificar Tabela1 e outros objetos de tabela
"""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

def explorar_estrutura_excel():
    """Explora a estrutura detalhada dos arquivos Excel"""
    
    print("🔍 EXPLORAÇÃO DETALHADA DOS ARQUIVOS EXCEL")
    print("=" * 60)
    
    arquivos = list(Path("data/raw").glob("*.xlsx"))
    
    if not arquivos:
        print("❌ Nenhum arquivo encontrado!")
        return
    
    for arquivo in arquivos:
        print(f"\n📁 ARQUIVO: {arquivo.name}")
        print("-" * 50)
        
        try:
            # 1. Análise com pandas (método tradicional)
            print("📊 1. Análise com Pandas:")
            
            # Listar todas as sheets
            sheets = pd.ExcelFile(arquivo).sheet_names
            print(f"   📋 Sheets encontradas: {sheets}")
            
            for sheet in sheets:
                try:
                    df = pd.read_excel(arquivo, sheet_name=sheet)
                    print(f"   📄 Sheet '{sheet}': {len(df)} linhas, {len(df.columns)} colunas")
                    
                    # Mostrar primeiras colunas
                    if len(df.columns) > 0:
                        colunas = list(df.columns)[:5]
                        print(f"      🏷️ Primeiras colunas: {colunas}")
                        
                        # Verificar se há dados
                        dados_nao_vazios = df.count().sum()
                        print(f"      📈 Células com dados: {dados_nao_vazios}")
                        
                except Exception as e:
                    print(f"   ❌ Erro ao ler sheet '{sheet}': {e}")
            
            # 2. Análise com openpyxl (para detectar tabelas)
            print(f"\n🔧 2. Análise com OpenPyXL (Tabelas):")
            
            wb = load_workbook(arquivo, read_only=True)
            
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                print(f"   📄 Sheet '{ws_name}':")
                
                # Verificar tabelas
                if hasattr(ws, 'tables') and ws.tables:
                    print(f"      📋 Tabelas encontradas: {len(ws.tables)}")
                    
                    for table_name, table in ws.tables.items():
                        print(f"      🎯 Tabela: '{table_name}'")
                        print(f"         📍 Range: {table.ref}")
                        
                        # Tentar ler a tabela especificamente
                        try:
                            # Converter range para coordenadas
                            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(table.ref)
                            
                            print(f"         📊 Dimensões: {max_row - min_row + 1} linhas, {max_col - min_col + 1} colunas")
                            
                            # Ler dados da tabela usando pandas com range específico
                            df_table = pd.read_excel(
                                arquivo, 
                                sheet_name=ws_name,
                                skiprows=min_row-1,
                                nrows=max_row-min_row+1,
                                usecols=range(min_col-1, max_col)
                            )
                            
                            if not df_table.empty:
                                print(f"         ✅ Dados carregados: {len(df_table)} linhas")
                                print(f"         🏷️ Colunas: {list(df_table.columns)[:3]}...")
                                
                                # Verificar se parece com dados de OS
                                colunas_os = ['nome', 'cliente', 'cpf', 'os', 'data', 'telefone']
                                colunas_encontradas = []
                                for col in df_table.columns:
                                    col_lower = str(col).lower()
                                    for termo_os in colunas_os:
                                        if termo_os in col_lower:
                                            colunas_encontradas.append(col)
                                            break
                                
                                if colunas_encontradas:
                                    print(f"         🎯 Colunas relevantes: {colunas_encontradas}")
                                else:
                                    print(f"         ⚠️ Nenhuma coluna de OS identificada")
                            
                        except Exception as e:
                            print(f"         ❌ Erro ao ler tabela: {e}")
                            
                else:
                    print(f"      ❌ Nenhuma tabela encontrada")
                    
                    # Verificar dados brutos na sheet
                    try:
                        # Verificar se há dados nas primeiras linhas
                        dados_raw = []
                        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                            dados_raw.append(row)
                        
                        if any(any(cell for cell in row if cell) for row in dados_raw):
                            print(f"      📊 Dados brutos encontrados nas primeiras 5 linhas")
                            
                            # Mostrar primeira linha não vazia
                            for i, row in enumerate(dados_raw, 1):
                                if any(cell for cell in row if cell):
                                    valores = [str(cell)[:15] if cell else "" for cell in row[:5]]
                                    print(f"         Linha {i}: {valores}")
                                    break
                        else:
                            print(f"      ❌ Nenhum dado encontrado")
                            
                    except Exception as e:
                        print(f"      ❌ Erro ao verificar dados brutos: {e}")
            
            wb.close()
            
        except Exception as e:
            print(f"❌ Erro ao analisar {arquivo.name}: {e}")

def testar_carregamento_tabela1():
    """Testa especificamente o carregamento da Tabela1"""
    
    print(f"\n🎯 TESTE ESPECÍFICO - CARREGAMENTO TABELA1")
    print("=" * 60)
    
    arquivos = list(Path("data/raw").glob("*.xlsx"))
    
    for arquivo in arquivos:
        print(f"\n📁 Testando: {arquivo.name}")
        
        try:
            # Método 1: Carregar sheet 0 diretamente
            df1 = pd.read_excel(arquivo, sheet_name=0)
            print(f"✅ Método 1 (sheet 0): {len(df1)} linhas, {len(df1.columns)} colunas")
            
            if not df1.empty:
                # Mostrar estrutura
                print(f"   🏷️ Colunas: {list(df1.columns)}")
                
                # Verificar se primeira linha são cabeçalhos válidos
                primeira_linha = df1.iloc[0] if len(df1) > 0 else None
                if primeira_linha is not None:
                    print(f"   📋 Primeira linha: {primeira_linha.tolist()[:3]}...")
                
                # Procurar por padrões de dados de OS
                colunas_texto = [str(col).lower() for col in df1.columns]
                termos_relevantes = ['nome', 'cliente', 'cpf', 'telefone', 'os', 'data']
                
                relevantes_encontradas = []
                for termo in termos_relevantes:
                    for col in colunas_texto:
                        if termo in col:
                            relevantes_encontradas.append(termo)
                            break
                
                if relevantes_encontradas:
                    print(f"   🎯 Termos relevantes: {relevantes_encontradas}")
                    print(f"   ✅ Este arquivo parece conter dados de OS!")
                else:
                    print(f"   ⚠️ Nenhum termo de OS identificado nas colunas")
                    
        except Exception as e:
            print(f"❌ Erro: {e}")

if __name__ == "__main__":
    explorar_estrutura_excel()
    testar_carregamento_tabela1()