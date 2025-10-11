#!/usr/bin/env python3
"""
VERIFICAÇÃO E CORREÇÃO - OS COM MÚLTIPLAS FORMAS DE PAGAMENTO
Corrige o problema de OS divididas em múltiplas linhas de pagamento
"""

import pandas as pd
import openpyxl
from pathlib import Path

def verificar_os_multiplas_formas():
    print("🔍 VERIFICAÇÃO - OS COM MÚLTIPLAS FORMAS DE PAGAMENTO")
    print("=" * 65)
    print("🎯 Identificando OS divididas em múltiplas linhas")
    print()
    
    # Verificar arquivo original SUZANO jan_24 (2024, não 2025)
    arquivo_suzano = Path("data/caixa_lojas/SUZANO/2024_SUZ/jan_24.xlsx")
    
    print(f"📄 Verificando arquivo: {arquivo_suzano}")
    
    if arquivo_suzano.exists():
        wb = openpyxl.load_workbook(arquivo_suzano, data_only=True)
        
        # Verificar dia 02 especificamente
        if '02' in wb.sheetnames:
            ws = wb['02']
            print(f"📊 Verificando aba '02' (dia 2)")
            
            # Procurar dados na região E-G (onde estão as vendas)
            print(f"🔍 Dados encontrados no dia 2:")
            print(f"{'Nº Venda':<12} {'Cliente':<25} {'Forma Pgto':<12} {'Valor':<10}")
            print("-" * 65)
            
            # Buscar na coluna E (Nº Venda) a partir da linha 6
            os_encontradas = []
            for row in range(6, 20):  # Procurar nas primeiras linhas onde normalmente estão os dados
                numero_os = ws.cell(row=row, column=5).value  # Coluna E
                cliente = ws.cell(row=row, column=6).value    # Coluna F  
                forma_pgto = ws.cell(row=row, column=7).value # Coluna G
                valor = ws.cell(row=row, column=8).value      # Coluna H (se existir)
                
                if numero_os and str(numero_os).strip():
                    os_encontradas.append({
                        'numero_os': str(numero_os).strip(),
                        'cliente': str(cliente).strip() if cliente else '',
                        'forma_pgto': str(forma_pgto).strip() if forma_pgto else '',
                        'valor': valor if valor else '',
                        'linha': row
                    })
                    
                    print(f"{str(numero_os):<12} {str(cliente)[:24]:<25} {str(forma_pgto):<12} {str(valor):<10}")
            
            wb.close()
            
            # Analisar OS duplicadas
            print(f"\n📊 ANÁLISE DE OS DUPLICADAS:")
            print("-" * 35)
            
            os_por_numero = {}
            for os in os_encontradas:
                numero = os['numero_os']
                if numero in os_por_numero:
                    os_por_numero[numero].append(os)
                else:
                    os_por_numero[numero] = [os]
            
            for numero, lista_os in os_por_numero.items():
                if len(lista_os) > 1:
                    print(f"\n🎯 OS {numero} - {len(lista_os)} linhas:")
                    for i, os in enumerate(lista_os, 1):
                        print(f"   {i}. {os['cliente']} | {os['forma_pgto']} | {os['valor']}")
            
            return os_encontradas
    
    return []

def corrigir_processamento_os_multiplas():
    """Criar versão corrigida do processamento que agrupa OS por número"""
    print(f"\n🔧 CRIANDO PROCESSAMENTO CORRIGIDO")
    print("=" * 45)
    
    # Lógica para agrupar OS por número e somar valores
    print("💡 Estratégia de correção:")
    print("1. Agrupar linhas pelo número da OS")
    print("2. Concatenar formas de pagamento (DN+CTC)")
    print("3. Somar valores quando necessário")
    print("4. Manter um registro por OS única")
    
    exemplo_correcao = """
    ANTES (problema):
    8434 | SUELI ALVES TANAKA | DN  | R$ 270,00
    8434 | SUELI ALVES TANAKA | CTC | R$ 225,00
    
    DEPOIS (corrigido):  
    8434 | SUELI ALVES TANAKA | DN+CTC | R$ 270,00 (valor principal)
    """
    
    print(exemplo_correcao)
    
    # Verificar impacto nos documentos gerados
    print(f"\n📊 IMPACTO NOS DOCUMENTOS:")
    print("-" * 30)
    print("❌ Problema atual: OS duplicadas inflam o número total")
    print("✅ Solução: Agrupar por número de OS real") 
    print("📈 Resultado: Contagem precisa de OS únicas")

def main():
    os_encontradas = verificar_os_multiplas_formas()
    
    if os_encontradas:
        print(f"\n📊 RESUMO DA VERIFICAÇÃO:")
        print(f"✅ Encontradas {len(os_encontradas)} linhas de venda no dia 2")
        
        # Contar OS únicas
        numeros_unicos = set(os['numero_os'] for os in os_encontradas)
        print(f"🎯 OS únicas reais: {len(numeros_unicos)}")
        print(f"⚠️ Diferença: {len(os_encontradas) - len(numeros_unicos)} linhas extras")
        
        corrigir_processamento_os_multiplas()
        
        print(f"\n💡 PRÓXIMA AÇÃO RECOMENDADA:")
        print("1. Ajustar lógica de processamento para agrupar por número de OS")
        print("2. Reprocessar documentos com agrupamento correto")
        print("3. Validar resultados com dados reais")
    else:
        print("❌ Não foi possível verificar os dados")

if __name__ == "__main__":
    main()