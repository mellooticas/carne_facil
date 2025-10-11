#!/usr/bin/env python3
"""Investigar estrutura dos arquivos Excel"""

import openpyxl
from pathlib import Path

def investigar_arquivo(caminho_arquivo):
    print(f"📄 INVESTIGANDO: {caminho_arquivo}")
    print("=" * 50)
    
    try:
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        print(f"📋 Total de abas: {len(wb.sheetnames)}")
        print(f"📋 Primeiras abas: {wb.sheetnames[:10]}")
        
        # Testar aba '02' que sabemos ter dados
        if '02' in wb.sheetnames:
            ws = wb['02']
            print(f"\n📊 ABA '02':")
            print(f"   Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
            
            # Procurar VEND
            vend_encontrada = False
            for row in range(1, min(30, ws.max_row)):
                for col in range(1, min(15, ws.max_column)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip().upper() == 'VEND':
                        print(f"   🎯 VEND encontrado em linha {row}, coluna {col}")
                        vend_encontrada = True
                        
                        # Mostrar dados após VEND
                        print(f"   📋 Dados após VEND:")
                        for r in range(row+1, min(row+8, ws.max_row)):
                            linha = []
                            for c in range(1, 8):
                                valor = ws.cell(row=r, column=c).value
                                linha.append(valor)
                            print(f"      Linha {r}: {linha}")
                        break
                if vend_encontrada:
                    break
            
            if not vend_encontrada:
                print("   ❌ VEND não encontrado")
                # Mostrar primeiras 10 linhas para entender a estrutura
                print("   📋 Primeiras 10 linhas:")
                for r in range(1, min(11, ws.max_row)):
                    linha = []
                    for c in range(1, 8):
                        valor = ws.cell(row=r, column=c).value
                        linha.append(valor)
                    print(f"      Linha {r}: {linha}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Erro: {e}")

# Testar arquivo que sabemos que funciona
arquivo_teste = Path("data/caixa_lojas/MAUA/jan_25.xlsx")
investigar_arquivo(arquivo_teste)