#!/usr/bin/env python3
"""
PROCESSADOR VENDAS REAL - ESTRUTURA CORRETA DESCOBERTA
Foca na região E5-H5 (Nº Venda, Cliente, Valor Venda) das sheets de dias
Objetivo: Capturar os R$ 5,7M reais com estrutura correta
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import re

class ProcessadorVendasReal:
    def __init__(self):
        self.pasta_dados = Path("data/caixa_lojas")
        self.pasta_resultados = Path("data/vendas_estrutura_real")
        self.pasta_resultados.mkdir(parents=True, exist_ok=True)
        
        self.lojas_ativas = ['MAUA', 'SUZANO', 'RIO_PEQUENO', 'PERUS', 'SUZANO2', 'SAO_MATEUS']
        
    def extrair_vendas_estrutura_real(self, arquivo_path, loja):
        """Extrai vendas da estrutura REAL descoberta: E5-H5 (Nº Venda, Cliente, Valor)"""
        print(f"   📄 Processando: {arquivo_path.name}")
        
        try:
            wb = openpyxl.load_workbook(arquivo_path, data_only=True)
            
            # Detectar ano e mês
            nome_arquivo = arquivo_path.name.lower()
            meses = {'jan': 'JAN', 'fev': 'FEV', 'mar': 'MAR', 'abr': 'ABR', 
                    'mai': 'MAI', 'jun': 'JUN', 'jul': 'JUL', 'ago': 'AGO',
                    'set': 'SET', 'out': 'OUT', 'nov': 'NOV', 'dez': 'DEZ'}
            
            mes_nome = 'UNKNOWN'
            for mes_abr, mes_nome_full in meses.items():
                if mes_abr in nome_arquivo:
                    mes_nome = mes_nome_full
                    break
            
            if '_25' in nome_arquivo or '2025' in nome_arquivo:
                ano = '2025'
            elif '_24' in nome_arquivo or '2024' in nome_arquivo:
                ano = '2024'
            elif '_23' in nome_arquivo or '2023' in nome_arquivo:
                ano = '2023'
            else:
                ano = '2024'
            
            todas_vendas = []
            
            # Processar cada dia (sheets numeradas)
            for dia in range(1, 32):
                dia_str = f"{dia:02d}"
                
                if dia_str in wb.sheetnames:
                    ws = wb[dia_str]
                    
                    # Verificar se existe cabeçalho correto na linha 5
                    col_e5 = ws.cell(row=5, column=5).value  # E5 - Nº Venda
                    col_f5 = ws.cell(row=5, column=6).value  # F5 - Cliente
                    col_h5 = ws.cell(row=5, column=8).value  # H5 - Valor Venda
                    
                    if col_e5 and 'venda' in str(col_e5).lower():
                        print(f"      📅 Dia {dia}: Cabeçalho encontrado - {col_e5} | {col_f5} | {col_h5}")
                        
                        # Extrair dados das vendas (linha 6 em diante)
                        vendas_dia = 0
                        for row in range(6, ws.max_row + 1):
                            numero_venda = ws.cell(row=row, column=5).value  # E - Nº Venda
                            cliente = ws.cell(row=row, column=6).value       # F - Cliente
                            valor_venda = ws.cell(row=row, column=8).value   # H - Valor Venda
                            forma_pgto = ws.cell(row=row, column=7).value    # G - Forma Pgto (se existir)
                            
                            # Só processar se tem número de venda
                            if numero_venda and str(numero_venda).strip():
                                numero_venda_str = str(numero_venda).strip()
                                cliente_nome = str(cliente).strip() if cliente else 'NAO_INFORMADO'
                                
                                # Tentar converter valor
                                valor_num = 0
                                if valor_venda:
                                    try:
                                        if isinstance(valor_venda, (int, float)):
                                            valor_num = float(valor_venda)
                                        else:
                                            # Remover formatação e converter
                                            valor_str = str(valor_venda).replace(',', '.').replace('R$', '').strip()
                                            valor_num = float(valor_str) if valor_str else 0
                                    except:
                                        valor_num = 0
                                
                                forma_pgto_str = str(forma_pgto).strip() if forma_pgto else 'NAO_INFORMADO'
                                
                                venda = {
                                    'loja': loja,
                                    'ano': ano,
                                    'mes': mes_nome,
                                    'dia': int(dia),
                                    'data_completa': f"{dia:02d}/{mes_nome.lower()[:3]}/{ano}",
                                    'numero_venda': numero_venda_str,
                                    'cliente': cliente_nome,
                                    'forma_pagamento': forma_pgto_str,
                                    'valor_venda': valor_num,
                                    'arquivo_origem': arquivo_path.name,
                                    'linha_origem': row
                                }
                                
                                todas_vendas.append(venda)
                                vendas_dia += 1
                        
                        if vendas_dia > 0:
                            print(f"         ✅ {vendas_dia} vendas extraídas")
                    else:
                        # Tentar buscar dados mesmo sem cabeçalho padrão
                        vendas_dia = 0
                        for row in range(6, min(40, ws.max_row + 1)):
                            numero_venda = ws.cell(row=row, column=5).value
                            cliente = ws.cell(row=row, column=6).value
                            valor_venda = ws.cell(row=row, column=8).value
                            
                            if numero_venda and str(numero_venda).strip() and str(numero_venda).isdigit():
                                numero_venda_str = str(numero_venda).strip()
                                cliente_nome = str(cliente).strip() if cliente else 'NAO_INFORMADO'
                                
                                valor_num = 0
                                if valor_venda:
                                    try:
                                        if isinstance(valor_venda, (int, float)):
                                            valor_num = float(valor_venda)
                                        else:
                                            valor_str = str(valor_venda).replace(',', '.').replace('R$', '').strip()
                                            valor_num = float(valor_str) if valor_str else 0
                                    except:
                                        valor_num = 0
                                
                                if valor_num > 0:  # Só incluir se tem valor
                                    venda = {
                                        'loja': loja,
                                        'ano': ano,
                                        'mes': mes_nome,
                                        'dia': int(dia),
                                        'data_completa': f"{dia:02d}/{mes_nome.lower()[:3]}/{ano}",
                                        'numero_venda': numero_venda_str,
                                        'cliente': cliente_nome,
                                        'forma_pagamento': 'NAO_INFORMADO',
                                        'valor_venda': valor_num,
                                        'arquivo_origem': arquivo_path.name,
                                        'linha_origem': row
                                    }
                                    
                                    todas_vendas.append(venda)
                                    vendas_dia += 1
                        
                        if vendas_dia > 0:
                            print(f"      📅 Dia {dia}: {vendas_dia} vendas extraídas (sem cabeçalho)")
            
            wb.close()
            
            print(f"      📊 TOTAL: {len(todas_vendas)} vendas extraídas")
            print(f"      💰 VALOR: R$ {sum(v['valor_venda'] for v in todas_vendas):,.2f}")
            
            return todas_vendas
            
        except Exception as e:
            print(f"      ❌ Erro: {e}")
            return []
    
    def processar_loja_estrutura_real(self, loja):
        """Processa loja com estrutura real descoberta"""
        print(f"\n💰 PROCESSANDO VENDAS REAIS: {loja}")
        print("=" * 60)
        
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            print(f"❌ Loja {loja} não encontrada")
            return None
        
        todas_vendas = []
        arquivos_processados = 0
        
        # Processar arquivos da raiz (2025)
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~'):
                vendas = self.extrair_vendas_estrutura_real(arquivo, loja)
                todas_vendas.extend(vendas)
                if vendas:
                    arquivos_processados += 1
        
        # Processar pastas de anos
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir() and any(ano in pasta.name for ano in ['2023', '2024', '2025']):
                for arquivo in pasta.glob("*.xlsx"):
                    if not arquivo.name.startswith('~'):
                        vendas = self.extrair_vendas_estrutura_real(arquivo, loja)
                        todas_vendas.extend(vendas)
                        if vendas:
                            arquivos_processados += 1
        
        if todas_vendas:
            # Criar DataFrame
            df = pd.DataFrame(todas_vendas)
            
            # Consolidar por OS única (agrupamento correto)
            print(f"\n🔧 CONSOLIDANDO OS ÚNICAS...")
            df['id_os_unico'] = df['loja'] + '_' + df['numero_venda'] + '_' + df['data_completa']
            
            # Agrupar por OS única
            df_consolidado = df.groupby('id_os_unico').agg({
                'loja': 'first',
                'ano': 'first',
                'mes': 'first',
                'dia': 'first',
                'data_completa': 'first',
                'numero_venda': 'first',
                'cliente': 'first',
                'forma_pagamento': lambda x: '+'.join(x.unique()),
                'valor_venda': 'max',  # Pegar o maior valor (evita duplicação)
                'arquivo_origem': 'first',
                'linha_origem': 'first'
            }).reset_index()
            
            antes = len(df)
            depois = len(df_consolidado)
            
            if antes != depois:
                print(f"   🧹 Consolidadas {antes - depois} vendas duplicadas")
            
            # Salvar resultados
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"VENDAS_REAIS_{loja}_{timestamp}.xlsx"
            caminho_arquivo = self.pasta_resultados / nome_arquivo
            
            df_consolidado.to_excel(caminho_arquivo, index=False)
            
            print(f"\n📊 RESULTADO VENDAS REAIS {loja}:")
            print(f"   ✅ Arquivos processados: {arquivos_processados}")
            print(f"   📈 Vendas únicas: {len(df_consolidado):,}")
            print(f"   💰 Valor total: R$ {df_consolidado['valor_venda'].sum():,.2f}")
            print(f"   📅 Período: {df_consolidado['data_completa'].min()} a {df_consolidado['data_completa'].max()}")
            print(f"   💾 Arquivo salvo: {nome_arquivo}")
            
            return {
                'loja': loja,
                'arquivo': nome_arquivo,
                'vendas_unicas': len(df_consolidado),
                'valor_total': df_consolidado['valor_venda'].sum(),
                'consolidacoes': antes - depois
            }
        
        return None
    
    def processar_todas_lojas_reais(self):
        """Processa todas as lojas com estrutura real"""
        print("💰 PROCESSAMENTO VENDAS REAIS - TODAS AS LOJAS")
        print("=" * 70)
        print("🎯 Estrutura descoberta: E5(Nº Venda) F5(Cliente) H5(Valor)")
        print()
        
        resultados = []
        
        for loja in self.lojas_ativas:
            resultado = self.processar_loja_estrutura_real(loja)
            if resultado:
                resultados.append(resultado)
        
        if resultados:
            print(f"\n🎉 PROCESSAMENTO VENDAS REAIS CONCLUÍDO!")
            print("=" * 50)
            
            total_vendas = sum(r['vendas_unicas'] for r in resultados)
            total_valor = sum(r['valor_total'] for r in resultados)
            total_consolidacoes = sum(r['consolidacoes'] for r in resultados)
            
            print(f"📊 RESULTADO FINAL VENDAS REAIS:")
            print(f"   🏢 Lojas processadas: {len(resultados)}")
            print(f"   📈 Vendas únicas totais: {total_vendas:,}")
            print(f"   💰 Valor total: R$ {total_valor:,.2f}")
            print(f"   🔧 Consolidações aplicadas: {total_consolidacoes}")
            
            # Verificar se chegamos próximo dos R$ 5,7M
            if total_valor >= 5000000:  # R$ 5M+
                print(f"   🎯 ✅ META ALCANÇADA! Valor próximo dos R$ 5,7M esperados!")
            else:
                print(f"   📊 Valor atual: {(total_valor/5700000)*100:.1f}% dos R$ 5,7M esperados")
            
            print(f"\n📄 ARQUIVOS GERADOS:")
            for resultado in resultados:
                print(f"   ✅ {resultado['loja']}: {resultado['arquivo']}")
                print(f"      📊 {resultado['vendas_unicas']:,} vendas | R$ {resultado['valor_total']:,.2f}")
        
        return resultados

def main():
    processador = ProcessadorVendasReal()
    
    print("💰 PROCESSADOR VENDAS REAIS - ESTRUTURA DESCOBERTA")
    print("=" * 65)
    print("🎯 Foco na estrutura real: E5(Nº Venda) F5(Cliente) H5(Valor)")
    print("🎯 Objetivo: Capturar os R$ 5,7M reais")
    print()
    print("1. Processar todas as lojas (estrutura real)")
    print("2. Processar loja específica")
    print("3. Sair")
    
    while True:
        escolha = input("\n👉 Escolha uma opção (1-3): ").strip()
        
        if escolha == "1":
            processador.processar_todas_lojas_reais()
            break
        elif escolha == "2":
            loja = input("Digite o nome da loja: ").strip().upper()
            if loja in processador.lojas_ativas:
                processador.processar_loja_estrutura_real(loja)
            else:
                print(f"❌ Loja {loja} não encontrada")
            break
        elif escolha == "3":
            print("👋 Saindo...")
            break
        else:
            print("❌ Opção inválida")

if __name__ == "__main__":
    main()