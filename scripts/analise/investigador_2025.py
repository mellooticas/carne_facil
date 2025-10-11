#!/usr/bin/env python3
"""
INVESTIGADOR 2025 - Verificar meses faltantes para completar R$ 5,7M
Análise focada nos dados de 2025 para identificar lacunas
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import calendar

class Investigador2025:
    def __init__(self):
        self.pasta_dados = Path("data/caixa_lojas")
        self.lojas_ativas = ['MAUA', 'SUZANO', 'RIO_PEQUENO', 'PERUS', 'SUZANO2', 'SAO_MATEUS']
        
        self.meses_completos = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 
                               'jul', 'ago', 'set', 'out', 'nov', 'dez']
        
        # Resultado atual por loja
        self.resultado_atual = {
            'MAUA': {'vendas': 4279, 'valor': 570853.41},
            'SUZANO': {'vendas': 8498, 'valor': 1947453.53},
            'RIO_PEQUENO': {'vendas': 4480, 'valor': 766814.20},
            'PERUS': {'vendas': 3930, 'valor': 439720.04},
            'SUZANO2': {'vendas': 2751, 'valor': 199746.98},
            'SAO_MATEUS': {'vendas': 2811, 'valor': 228937.73}
        }
    
    def verificar_arquivos_2025_loja(self, loja):
        """Verifica quais arquivos de 2025 existem para uma loja"""
        print(f"\n🔍 INVESTIGANDO 2025 - {loja}")
        print("=" * 50)
        
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            print(f"❌ Loja {loja} não encontrada")
            return None
        
        arquivos_2025 = []
        meses_encontrados = []
        
        # Buscar arquivos de 2025 na raiz
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~'):
                nome = arquivo.name.lower()
                if '_25' in nome or '2025' in nome:
                    arquivos_2025.append(arquivo)
                    
                    # Identificar mês
                    for mes in self.meses_completos:
                        if mes in nome:
                            meses_encontrados.append(mes.upper())
                            break
        
        # Buscar em pastas de 2025
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir() and '2025' in pasta.name:
                for arquivo in pasta.glob("*.xlsx"):
                    if not arquivo.name.startswith('~'):
                        arquivos_2025.append(arquivo)
                        
                        nome = arquivo.name.lower()
                        for mes in self.meses_completos:
                            if mes in nome:
                                meses_encontrados.append(mes.upper())
                                break
        
        meses_encontrados = list(set(meses_encontrados))  # Remover duplicatas
        meses_faltantes = [mes.upper() for mes in self.meses_completos if mes.upper() not in meses_encontrados]
        
        print(f"📅 ARQUIVOS 2025 ENCONTRADOS: {len(arquivos_2025)}")
        for arquivo in arquivos_2025:
            print(f"   📄 {arquivo.name}")
        
        print(f"\n📊 ANÁLISE MENSAL 2025:")
        print(f"   ✅ Meses encontrados ({len(meses_encontrados)}): {', '.join(meses_encontrados)}")
        if meses_faltantes:
            print(f"   ❌ Meses faltantes ({len(meses_faltantes)}): {', '.join(meses_faltantes)}")
        else:
            print(f"   🎉 Todos os meses de 2025 presentes!")
        
        return {
            'loja': loja,
            'arquivos_2025': len(arquivos_2025),
            'meses_encontrados': meses_encontrados,
            'meses_faltantes': meses_faltantes,
            'cobertura_percentual': (len(meses_encontrados) / 12) * 100
        }
    
    def analisar_vendas_2025_rapido(self, loja):
        """Análise rápida das vendas de 2025 de uma loja"""
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            return None
        
        vendas_2025 = []
        valor_total_2025 = 0
        
        # Processar arquivos de 2025
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~'):
                nome = arquivo.name.lower()
                if '_25' in nome or '2025' in nome:
                    try:
                        wb = openpyxl.load_workbook(arquivo, data_only=True)
                        
                        # Processar dias
                        for dia in range(1, 32):
                            dia_str = f"{dia:02d}"
                            
                            if dia_str in wb.sheetnames:
                                ws = wb[dia_str]
                                
                                # Extrair vendas rápido (só valores)
                                for row in range(6, min(40, ws.max_row + 1)):
                                    numero_venda = ws.cell(row=row, column=5).value
                                    valor_venda = ws.cell(row=row, column=8).value
                                    
                                    if numero_venda and str(numero_venda).strip():
                                        if valor_venda:
                                            try:
                                                if isinstance(valor_venda, (int, float)):
                                                    valor_num = float(valor_venda)
                                                else:
                                                    valor_str = str(valor_venda).replace(',', '.').replace('R$', '').strip()
                                                    valor_num = float(valor_str) if valor_str else 0
                                                
                                                if valor_num > 0:
                                                    vendas_2025.append(valor_num)
                                                    valor_total_2025 += valor_num
                                            except:
                                                pass
                        
                        wb.close()
                        
                    except Exception as e:
                        print(f"      ⚠️ Erro em {arquivo.name}: {e}")
        
        return {
            'vendas_2025': len(vendas_2025),
            'valor_2025': valor_total_2025
        }
    
    def investigar_todas_lojas_2025(self):
        """Investiga dados de 2025 de todas as lojas"""
        print("🔍 INVESTIGAÇÃO COMPLETA 2025 - BUSCA PELOS R$ 1,5M FALTANTES")
        print("=" * 70)
        print("🎯 Meta: Encontrar dados faltantes para completar R$ 5,7M")
        print()
        
        resultados = []
        total_valor_2025 = 0
        
        for loja in self.lojas_ativas:
            resultado = self.verificar_arquivos_2025_loja(loja)
            
            if resultado:
                # Análise rápida de vendas
                vendas_2025 = self.analisar_vendas_2025_rapido(loja)
                
                if vendas_2025:
                    resultado['vendas_2025'] = vendas_2025['vendas_2025']
                    resultado['valor_2025'] = vendas_2025['valor_2025']
                    total_valor_2025 += vendas_2025['valor_2025']
                    
                    print(f"\n💰 VENDAS 2025 {loja}:")
                    print(f"   📈 Vendas: {vendas_2025['vendas_2025']:,}")
                    print(f"   💵 Valor: R$ {vendas_2025['valor_2025']:,.2f}")
                    
                    # Comparar com resultado atual (que pode incluir dados de anos anteriores)
                    valor_atual = self.resultado_atual[loja]['valor']
                    percentual_2025 = (vendas_2025['valor_2025'] / valor_atual) * 100 if valor_atual > 0 else 0
                    print(f"   📊 Representa {percentual_2025:.1f}% do valor atual da loja")
                else:
                    resultado['vendas_2025'] = 0
                    resultado['valor_2025'] = 0
                
                resultados.append(resultado)
        
        # Análise consolidada
        print(f"\n🎯 ANÁLISE CONSOLIDADA 2025")
        print("=" * 50)
        
        total_cobertura = sum(r['cobertura_percentual'] for r in resultados) / len(resultados)
        total_meses_faltantes = sum(len(r['meses_faltantes']) for r in resultados)
        
        print(f"📊 COBERTURA MÉDIA 2025: {total_cobertura:.1f}%")
        print(f"📅 TOTAL MESES FALTANTES: {total_meses_faltantes}")
        print(f"💰 VALOR TOTAL 2025: R$ {total_valor_2025:,.2f}")
        
        # Estimativa do valor faltante
        if total_cobertura < 100:
            estimativa_valor_completo = total_valor_2025 * (100 / total_cobertura) if total_cobertura > 0 else 0
            valor_estimado_faltante = estimativa_valor_completo - total_valor_2025
            
            print(f"📈 ESTIMATIVA VALOR COMPLETO 2025: R$ {estimativa_valor_completo:,.2f}")
            print(f"🔍 VALOR ESTIMADO FALTANTE: R$ {valor_estimado_faltante:,.2f}")
            
            # Verificar se isso pode completar os R$ 5,7M
            valor_atual_total = 4153525.89  # Do resultado anterior
            valor_projetado_total = valor_atual_total + valor_estimado_faltante
            
            print(f"\n🎯 PROJEÇÃO PARA META R$ 5,7M:")
            print(f"   📊 Valor atual: R$ {valor_atual_total:,.2f}")
            print(f"   ➕ Estimativa faltante: R$ {valor_estimado_faltante:,.2f}")
            print(f"   🎯 Total projetado: R$ {valor_projetado_total:,.2f}")
            print(f"   📈 Percentual da meta: {(valor_projetado_total/5700000)*100:.1f}%")
            
            if valor_projetado_total >= 5700000:
                print(f"   🎉 META ALCANÇÁVEL COM DADOS COMPLETOS!")
            else:
                diferenca = 5700000 - valor_projetado_total
                print(f"   📊 Ainda faltariam: R$ {diferenca:,.2f}")
        
        print(f"\n📋 RESUMO POR LOJA:")
        for resultado in resultados:
            print(f"   🏢 {resultado['loja']}:")
            print(f"      📅 Cobertura: {resultado['cobertura_percentual']:.1f}%")
            print(f"      💰 Valor 2025: R$ {resultado.get('valor_2025', 0):,.2f}")
            if resultado['meses_faltantes']:
                print(f"      ❌ Faltantes: {', '.join(resultado['meses_faltantes'])}")
        
        return resultados

def main():
    investigador = Investigador2025()
    
    print("🔍 INVESTIGADOR 2025 - BUSCA PELOS R$ 1,5M FALTANTES")
    print("=" * 60)
    print("🎯 Objetivo: Verificar dados de 2025 para completar R$ 5,7M")
    print()
    print("1. Investigar todas as lojas (2025)")
    print("2. Investigar loja específica")
    print("3. Sair")
    
    while True:
        escolha = input("\n👉 Escolha uma opção (1-3): ").strip()
        
        if escolha == "1":
            investigador.investigar_todas_lojas_2025()
            break
        elif escolha == "2":
            loja = input("Digite o nome da loja: ").strip().upper()
            if loja in investigador.lojas_ativas:
                investigador.verificar_arquivos_2025_loja(loja)
                vendas = investigador.analisar_vendas_2025_rapido(loja)
                if vendas:
                    print(f"\n💰 VENDAS 2025: {vendas['vendas_2025']:,} | R$ {vendas['valor_2025']:,.2f}")
            else:
                print(f"❌ Loja {loja} não encontrada")
            break
        elif escolha == "3":
            print("👋 Saindo...")
            break
        else:
            print("❌ Opção inválida")

if __name__ == "__main__":
    main()