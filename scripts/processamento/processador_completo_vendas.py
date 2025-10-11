#!/usr/bin/env python3
"""
PROCESSADOR COMPLETO DE VENDAS - DOCUMENTOS FINAIS
Gera documentos consolidados com todos os dados de cada loja
"""

import pandas as pd
from pathlib import Path
import openpyxl
from datetime import datetime
import re

class ProcessadorCompletoVendas:
    def __init__(self):
        self.pasta_dados = Path("data/caixa_lojas")
        self.pasta_finais = Path("data/documentos_finais")
        self.pasta_finais.mkdir(parents=True, exist_ok=True)
        
        # Configurações de processamento
        self.colunas_padrao = ['loja', 'data', 'numero_venda', 'cliente', 'forma_pgto', 'valor_venda', 'entrada']
        
        # Lojas conhecidas (incluindo São Mateus para análise histórica)
        self.lojas_ativas = ['MAUA', 'SUZANO', 'RIO_PEQUENO', 'PERUS', 'SUZANO2', 'SAO_MATEUS']
        
    def detectar_estrutura_loja(self, loja):
        """Detecta todos os arquivos disponíveis para uma loja"""
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            return None
        
        estrutura = {
            'arquivos_2025': [],  # Arquivos direto na pasta
            'arquivos_2024': [],  # Arquivos na pasta 2024_XXX
            'outras_pastas': [],
            'total_arquivos': 0
        }
        
        # Arquivos diretos (geralmente 2025)
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~'):  # Ignorar arquivos temporários
                estrutura['arquivos_2025'].append(arquivo)
                estrutura['total_arquivos'] += 1
        
        # Pastas com anos
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir():
                nome_pasta = pasta.name.lower()
                
                if '2024' in nome_pasta:
                    arquivos_2024 = list(pasta.glob("*.xlsx"))
                    estrutura['arquivos_2024'].extend(arquivos_2024)
                    estrutura['total_arquivos'] += len(arquivos_2024)
                
                elif any(ano in nome_pasta for ano in ['2023', '2025']):
                    arquivos_outros = list(pasta.glob("*.xlsx"))
                    estrutura['outras_pastas'].append({
                        'pasta': pasta.name,
                        'arquivos': arquivos_outros
                    })
                    estrutura['total_arquivos'] += len(arquivos_outros)
        
        return estrutura
    
    def extrair_vendas_arquivo(self, arquivo_path, loja):
        """Extrai vendas de um arquivo específico usando a lógica do sistema universal"""
        print(f"   📄 Processando: {arquivo_path.name}")
        
        try:
            # Carregar workbook
            wb = openpyxl.load_workbook(arquivo_path, data_only=True)
            vendas_consolidadas = []
            
            # Detectar padrão do nome do arquivo para definir mês/ano
            nome_arquivo = arquivo_path.name.lower()
            
            # Extrair ano e mês do nome do arquivo
            if 'jan' in nome_arquivo:
                mes_nome = 'JAN'
            elif 'fev' in nome_arquivo:
                mes_nome = 'FEV'
            elif 'mar' in nome_arquivo:
                mes_nome = 'MAR'
            elif 'abr' in nome_arquivo:
                mes_nome = 'ABR'
            elif 'mai' in nome_arquivo:
                mes_nome = 'MAI'
            elif 'jun' in nome_arquivo:
                mes_nome = 'JUN'
            elif 'jul' in nome_arquivo:
                mes_nome = 'JUL'
            elif 'ago' in nome_arquivo:
                mes_nome = 'AGO'
            elif 'set' in nome_arquivo:
                mes_nome = 'SET'
            elif 'out' in nome_arquivo:
                mes_nome = 'OUT'
            elif 'nov' in nome_arquivo:
                mes_nome = 'NOV'
            elif 'dez' in nome_arquivo:
                mes_nome = 'DEZ'
            else:
                mes_nome = 'UNKNOWN'
            
            # Detectar ano
            if '_25' in nome_arquivo or '2025' in nome_arquivo:
                ano = '2025'
            elif '_24' in nome_arquivo or '2024' in nome_arquivo:
                ano = '2024'
            elif '_23' in nome_arquivo or '2023' in nome_arquivo:
                ano = '2023'
            else:
                ano = '2024'  # Default
            
            # Processar dias (assumindo padrão de 1 a 31)
            for dia in range(1, 32):
                dia_str = f"{dia:02d}"
                
                try:
                    # Tentar acessar a aba do dia
                    if dia_str in wb.sheetnames:
                        ws = wb[dia_str]
                        
                        # Procurar cabeçalho "Nº Venda" na coluna E (5)
                        header_encontrado = False
                        start_row = None
                        
                        for row_num in range(1, min(20, ws.max_row)):
                            cell_value = ws.cell(row=row_num, column=5).value  # Coluna E
                            if cell_value and 'Nº Venda' in str(cell_value):
                                header_encontrado = True
                                start_row = row_num + 1
                                break
                        
                        if header_encontrado and start_row:
                            # Extrair dados da tabela (colunas E, F, G = nº_venda, cliente, forma_pgto)
                            vendas_dia = []
                            for row_num in range(start_row, ws.max_row + 1):
                                # Dados nas colunas E=5, F=6, G=7
                                numero_venda = ws.cell(row=row_num, column=5).value
                                cliente = ws.cell(row=row_num, column=6).value
                                forma_pgto = ws.cell(row=row_num, column=7).value
                                
                                # Verificar se há dados válidos
                                if numero_venda is not None and str(numero_venda).strip():
                                    # Buscar valor da venda na coluna C (3) - linha "Vendas"
                                    valor_venda = 0
                                    entrada = 0
                                    
                                    # Procurar valor nas primeiras linhas da planilha
                                    for search_row in range(1, 15):
                                        label = ws.cell(row=search_row, column=2).value
                                        if label and 'Vendas' in str(label):
                                            valor_venda = ws.cell(row=search_row, column=3).value or 0
                                        elif label and 'Entrada' in str(label):
                                            entrada = ws.cell(row=search_row, column=3).value or 0
                                    
                                    venda_dict = {
                                        'loja': loja,
                                        'data': f"{dia:02d}/{mes_nome.lower()[:3]}/{ano}",
                                        'numero_venda': numero_venda,
                                        'cliente': cliente or 'N/A',
                                        'forma_pgto': forma_pgto or 'N/A',
                                        'valor_venda': valor_venda,
                                        'entrada': entrada
                                    }
                                    vendas_consolidadas.append(venda_dict)
                                    vendas_dia.append(venda_dict)
                                else:
                                    # Para de processar se encontrou linha vazia
                                    break
                
                except Exception as e:
                    # Dia não existe ou erro, continuar
                    continue
            
            wb.close()
            print(f"      ✅ {len(vendas_consolidadas)} vendas extraídas")
            return vendas_consolidadas
            
        except Exception as e:
            print(f"      ❌ Erro: {e}")
            return []
    
    def processar_loja_completa(self, loja):
        """Processa todos os dados de uma loja e gera documento final"""
        print(f"\n🏢 PROCESSANDO LOJA COMPLETA: {loja}")
        print("=" * 60)
        
        # Detectar estrutura
        estrutura = self.detectar_estrutura_loja(loja)
        
        if not estrutura:
            print(f"❌ Loja {loja} não encontrada")
            return None
        
        print(f"📊 Estrutura detectada:")
        print(f"   📄 Arquivos 2025: {len(estrutura['arquivos_2025'])}")
        print(f"   📄 Arquivos 2024: {len(estrutura['arquivos_2024'])}")
        print(f"   📁 Outras pastas: {len(estrutura['outras_pastas'])}")
        print(f"   🎯 Total: {estrutura['total_arquivos']} arquivos")
        
        # Processar todos os arquivos
        todas_vendas = []
        arquivos_processados = 0
        arquivos_erro = 0
        
        # Processar arquivos 2025
        print(f"\n📥 Processando arquivos 2025...")
        for arquivo in estrutura['arquivos_2025']:
            vendas = self.extrair_vendas_arquivo(arquivo, loja)
            todas_vendas.extend(vendas)
            if vendas:
                arquivos_processados += 1
            else:
                arquivos_erro += 1
        
        # Processar arquivos 2024
        print(f"\n📥 Processando arquivos 2024...")
        for arquivo in estrutura['arquivos_2024']:
            vendas = self.extrair_vendas_arquivo(arquivo, loja)
            todas_vendas.extend(vendas)
            if vendas:
                arquivos_processados += 1
            else:
                arquivos_erro += 1
        
        # Processar outras pastas se necessário
        for pasta_info in estrutura['outras_pastas']:
            print(f"\n📥 Processando pasta {pasta_info['pasta']}...")
            for arquivo in pasta_info['arquivos']:
                vendas = self.extrair_vendas_arquivo(arquivo, loja)
                todas_vendas.extend(vendas)
                if vendas:
                    arquivos_processados += 1
                else:
                    arquivos_erro += 1
        
        # Criar DataFrame e salvar
        if todas_vendas:
            df = pd.DataFrame(todas_vendas)
            
            # Limpeza e padronização
            df = self.limpar_dados(df)
            
            # Salvar documento final
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"VENDAS_COMPLETAS_{loja}_{timestamp}.xlsx"
            caminho_final = self.pasta_finais / nome_arquivo
            
            df.to_excel(caminho_final, index=False)
            
            # Relatório
            print(f"\n📊 RESULTADO DO PROCESSAMENTO:")
            print(f"   ✅ Arquivos processados: {arquivos_processados}")
            print(f"   ❌ Arquivos com erro: {arquivos_erro}")
            print(f"   📈 Total de vendas: {len(df):,}")
            print(f"   💰 Valor total: R$ {df['valor_venda'].sum():,.2f}")
            print(f"   📅 Período: {df['data'].min()} a {df['data'].max()}")
            print(f"   💾 Arquivo salvo: {nome_arquivo}")
            
            return {
                'loja': loja,
                'arquivo': nome_arquivo,
                'vendas': len(df),
                'valor_total': df['valor_venda'].sum(),
                'periodo': f"{df['data'].min()} a {df['data'].max()}",
                'arquivos_processados': arquivos_processados
            }
        
        else:
            print(f"❌ Nenhuma venda encontrada para {loja}")
            return None
    
    def limpar_dados(self, df):
        """Limpa e padroniza os dados"""
        print(f"   🧹 Limpando dados...")
        
        # Converter valores para numérico
        df['valor_venda'] = pd.to_numeric(df['valor_venda'], errors='coerce').fillna(0)
        df['entrada'] = pd.to_numeric(df['entrada'], errors='coerce').fillna(0)
        
        # Limpar strings
        colunas_texto = ['cliente', 'forma_pgto']
        for col in colunas_texto:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
        
        # Remover linhas sem dados importantes
        df = df.dropna(subset=['numero_venda'])
        
        # Ordenar por data
        df = df.sort_values(['data', 'numero_venda'])
        
        print(f"      ✅ {len(df)} registros após limpeza")
        return df
    
    def processar_todas_lojas(self):
        """Processa todas as lojas ativas"""
        print("🏪 PROCESSAMENTO COMPLETO - TODAS AS LOJAS")
        print("=" * 70)
        
        resultados = []
        
        for loja in self.lojas_ativas:
            resultado = self.processar_loja_completa(loja)
            if resultado:
                resultados.append(resultado)
        
        # Relatório consolidado
        if resultados:
            print(f"\n🎉 PROCESSAMENTO CONCLUÍDO!")
            print("=" * 40)
            
            total_vendas = sum(r['vendas'] for r in resultados)
            total_valor = sum(r['valor_total'] for r in resultados)
            
            print(f"📊 RESUMO GERAL:")
            print(f"   🏢 Lojas processadas: {len(resultados)}")
            print(f"   📈 Total de vendas: {total_vendas:,}")
            print(f"   💰 Valor total: R$ {total_valor:,.2f}")
            
            print(f"\n📄 ARQUIVOS GERADOS:")
            for resultado in resultados:
                print(f"   ✅ {resultado['loja']}: {resultado['arquivo']}")
        
        return resultados

def main():
    processador = ProcessadorCompletoVendas()
    
    print("🏪 PROCESSADOR COMPLETO DE VENDAS")
    print("=" * 50)
    print("📋 Gera documentos finais consolidados por loja")
    print()
    print("1. Processar todas as lojas")
    print("2. Processar loja específica")
    print("3. Sair")
    
    while True:
        escolha = input("\n👉 Escolha uma opção (1-3): ").strip()
        
        if escolha == "1":
            processador.processar_todas_lojas()
            break
        elif escolha == "2":
            loja = input("Digite o nome da loja (MAUA, SUZANO, etc.): ").strip().upper()
            if loja in processador.lojas_ativas:
                processador.processar_loja_completa(loja)
            else:
                print(f"❌ Loja {loja} não encontrada")
            break
        elif escolha == "3":
            print("👋 Saindo...")
            break
        else:
            print("❌ Opção inválida. Tente novamente.")

if __name__ == "__main__":
    main()