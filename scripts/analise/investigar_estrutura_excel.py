#!/usr/bin/env python3
"""
INVESTIGADOR DE ESTRUTURA EXCEL - Descobrir formato real dos dados
"""

import pandas as pd
import openpyxl
from pathlib import Path

def investigar_arquivo_excel(arquivo_path):
    """Investiga a estrutura completa de um arquivo Excel"""
    print(f"\n🔍 INVESTIGANDO: {arquivo_path.name}")
    print("=" * 60)
    
    try:
        wb = openpyxl.load_workbook(arquivo_path, data_only=True)
        
        print(f"📋 SHEETS DISPONÍVEIS: {len(wb.sheetnames)}")
        for i, sheet_name in enumerate(wb.sheetnames):
            print(f"   {i+1:2d}. {sheet_name}")
        
        # Investigar algumas sheets específicas
        sheets_interessantes = []
        
        # Buscar por sheets de dias
        for sheet_name in wb.sheetnames:
            if sheet_name.isdigit() or any(d in sheet_name for d in ['01', '02', '03', '04', '05']):
                sheets_interessantes.append(sheet_name)
        
        # Pegar primeiras 3 sheets de dias
        if sheets_interessantes:
            print(f"\n📅 INVESTIGANDO SHEETS DE DIAS:")
            
            for sheet_name in sheets_interessantes[:3]:
                print(f"\n   🗓️ SHEET: {sheet_name}")
                ws = wb[sheet_name]
                
                print(f"      📏 Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
                
                # Buscar por texto "vend" em toda a sheet
                print(f"      🔍 Buscando 'vend' na sheet...")
                vend_encontrados = []
                
                for row in range(1, min(50, ws.max_row + 1)):
                    for col in range(1, min(20, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            if 'vend' in cell_value.lower():
                                coord = ws.cell(row=row, column=col).coordinate
                                vend_encontrados.append(f"{coord}: {cell_value}")
                
                if vend_encontrados:
                    print(f"      ✅ Encontrados {len(vend_encontrados)} textos com 'vend':")
                    for vend in vend_encontrados[:5]:  # Mostrar só os primeiros 5
                        print(f"         📌 {vend}")
                else:
                    print(f"      ❌ Nenhum texto com 'vend' encontrado")
                
                # Mostrar amostra das primeiras linhas e colunas
                print(f"      📊 AMOSTRA DOS DADOS (primeiras 10 linhas x 10 colunas):")
                for row in range(1, min(11, ws.max_row + 1)):
                    linha = []
                    for col in range(1, min(11, ws.max_column + 1)):
                        valor = ws.cell(row=row, column=col).value
                        if valor is not None:
                            valor_str = str(valor)[:15]  # Limitar tamanho
                            linha.append(valor_str)
                        else:
                            linha.append("")
                    
                    if any(linha):  # Só mostrar linhas que têm dados
                        print(f"         {row:2d}: {' | '.join(f'{v:15s}' for v in linha[:6])}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Erro ao investigar arquivo: {e}")

def main():
    pasta_dados = Path("data/caixa_lojas")
    
    print("🔍 INVESTIGADOR DE ESTRUTURA EXCEL")
    print("=" * 50)
    print("🎯 Descobrir formato real da tabela de vendas")
    
    # Buscar arquivo para investigar
    loja = "SUZANO"  # Vamos começar com SUZANO
    pasta_loja = pasta_dados / loja
    
    if pasta_loja.exists():
        # Pegar primeiro arquivo .xlsx encontrado
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~'):
                investigar_arquivo_excel(arquivo)
                break
        
        # Pegar também um arquivo de subpasta
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir() and '2024' in pasta.name:
                for arquivo in pasta.glob("*.xlsx"):
                    if not arquivo.name.startswith('~'):
                        investigar_arquivo_excel(arquivo)
                        break
                break

if __name__ == "__main__":
    main()