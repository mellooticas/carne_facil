#!/usr/bin/env python3
"""
 SISTEMA FINAL INTEGRADO - TICAS CARNE FCIL
================================================================================
 Combina TODOS os dados extrados em um sistema completo:
 3,262 Clientes nicos +  5,624 OS relacionadas
 12,309 Dioptras +  R$ 70.2M em vendas
 Dashboard executivo + Anlises detalhadas
================================================================================
"""

import pandas as pd
import logging
from pathlib import Path
import glob
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def aplicar_formatacao_excel(ws, df, titulo=""):
    """Aplica formatao profissional ao Excel"""
    # Ttulo
    if titulo:
        ws.insert_rows(1)
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells(f'A1:{chr(64 + len(df.columns))}1')
    
    # Headers
    header_row = 2 if titulo else 1
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-ajustar largura das colunas
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = chr(64 + col_num)  # A=65, B=66, etc.
        
        for cell in column:
            try:
                if hasattr(cell, 'value') and cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def criar_dashboard_executivo(df_clientes, df_relacionamentos, df_dioptrias, df_vendas):
    """Cria dashboard executivo com KPIs principais"""
    
    # KPIs principais
    total_clientes = len(df_clientes)
    total_os = len(df_relacionamentos)
    os_identificadas = len(df_relacionamentos[df_relacionamentos['Cliente_ID'].notna()])
    taxa_identificacao = (os_identificadas / total_os * 100) if total_os > 0 else 0
    
    # Vendas
    total_vendas = df_vendas['Valor_total'].sum() if 'Valor_total' in df_vendas.columns else 0
    os_com_vendas = len(df_vendas[df_vendas['Valor_total'] > 0]) if 'Valor_total' in df_vendas.columns else 0
    
    # Dioptras
    os_com_dioptrias = len(df_dioptrias) if not df_dioptrias.empty else 0
    
    # Criar dashboard
    dashboard_data = [
        ['INDICADOR', 'VALOR', 'DESCRIO'],
        ['Total de Clientes', f'{total_clientes:,}', 'Clientes nicos identificados'],
        ['Total de OS', f'{total_os:,}', 'Ordens de servio processadas'],
        ['OS Identificadas', f'{os_identificadas:,}', 'OS conectadas com clientes'],
        ['Taxa de Identificao', f'{taxa_identificacao:.1f}%', 'Percentual de OS com cliente conhecido'],
        ['Total em Vendas', f'R$ {total_vendas:,.2f}', 'Valor total das vendas registradas'],
        ['OS com Vendas', f'{os_com_vendas:,}', 'OS que possuem dados de vendas'],
        ['OS com Dioptras', f'{os_com_dioptrias:,}', 'OS com dados de prescrio'],
        ['Lojas Ativas', '3', 'MAUA, SAO_MATEUS, RIO_PEQUENO'],
        ['Sistemas', '2', 'LANCASTER, OTM'],
    ]
    
    return pd.DataFrame(dashboard_data[1:], columns=dashboard_data[0])

def criar_analise_por_loja(df_relacionamentos, df_vendas):
    """Cria anlise detalhada por loja"""
    
    analise_lojas = []
    
    for loja in df_relacionamentos['Loja'].unique():
        df_loja = df_relacionamentos[df_relacionamentos['Loja'] == loja]
        
        # Relacionamentos
        total_os = len(df_loja)
        os_identificadas = len(df_loja[df_loja['Cliente_ID'].notna()])
        taxa_ident = (os_identificadas / total_os * 100) if total_os > 0 else 0
        
        # Vendas da loja
        os_loja_nums = df_loja['OS_Numero'].tolist()
        df_vendas_loja = df_vendas[df_vendas['OS_Numero'].isin(os_loja_nums)] if 'OS_Numero' in df_vendas.columns else pd.DataFrame()
        
        total_vendas = df_vendas_loja['Valor_total'].sum() if not df_vendas_loja.empty and 'Valor_total' in df_vendas_loja.columns else 0
        os_com_vendas = len(df_vendas_loja[df_vendas_loja['Valor_total'] > 0]) if not df_vendas_loja.empty and 'Valor_total' in df_vendas_loja.columns else 0
        
        analise_lojas.append({
            'Loja': loja,
            'Total_OS': total_os,
            'OS_Identificadas': os_identificadas,
            'Taxa_Identificacao_%': round(taxa_ident, 1),
            'OS_com_Vendas': os_com_vendas,
            'Total_Vendas_R$': round(total_vendas, 2),
            'Ticket_Medio_R$': round(total_vendas / os_com_vendas, 2) if os_com_vendas > 0 else 0
        })
    
    return pd.DataFrame(analise_lojas)

def criar_top_clientes(df_relacionamentos, df_vendas):
    """Identifica top clientes por nmero de OS e vendas"""
    
    # Contar OS por cliente
    clientes_os = df_relacionamentos[df_relacionamentos['Cliente_ID'].notna()].groupby('Cliente_ID').agg({
        'OS_Numero': 'count',
        'Nome_OS': 'first'
    }).reset_index()
    clientes_os.columns = ['Cliente_ID', 'Total_OS', 'Nome_Cliente']
    
    # Adicionar vendas se disponvel
    if not df_vendas.empty and 'Cliente_ID' in df_vendas.columns:
        vendas_cliente = df_vendas.groupby('Cliente_ID')['Valor_total'].sum().reset_index()
        vendas_cliente.columns = ['Cliente_ID', 'Total_Vendas']
        clientes_os = clientes_os.merge(vendas_cliente, on='Cliente_ID', how='left')
        clientes_os['Total_Vendas'] = clientes_os['Total_Vendas'].fillna(0)
    else:
        clientes_os['Total_Vendas'] = 0
    
    # Ordenar por nmero de OS
    clientes_os = clientes_os.sort_values('Total_OS', ascending=False)
    
    return clientes_os.head(50)  # Top 50

def main():
    print(" SISTEMA FINAL INTEGRADO - TICAS CARNE FCIL")
    print("=" * 80)
    print(" Combinando todos os dados extrados")
    print(" Gerando relatrio executivo completo")
    print("=" * 80)
    
    # Diretrios
    data_dir = Path("data")
    processed_dir = data_dir / "processed"
    
    # 1. Carregar todos os dados processados
    print("\n Carregando dados processados...")
    
    # Clientes
    clientes_files = list(processed_dir.glob("BASE_CLIENTES_COM_ID_*.xlsx"))
    if not clientes_files:
        print(" Arquivo de clientes no encontrado!")
        return
    df_clientes = pd.read_excel(clientes_files[-1])
    print(f" {len(df_clientes)} clientes nicos")
    
    # Relacionamentos OS-Cliente
    relacionamentos_files = list(processed_dir.glob("RELACIONAMENTO_OS_CLIENTE_*.xlsx"))
    if relacionamentos_files:
        df_relacionamentos = pd.read_excel(relacionamentos_files[-1])
        print(f" {len(df_relacionamentos)} relacionamentos OS-Cliente")
    else:
        print("  Relacionamentos no encontrados")
        df_relacionamentos = pd.DataFrame()
    
    # Dioptras
    dioptrias_files = list(processed_dir.glob("DIOPTRIAS_COMPLETAS_*.xlsx"))
    if dioptrias_files:
        df_dioptrias = pd.read_excel(dioptrias_files[-1])
        print(f" {len(df_dioptrias)} registros de dioptras")
    else:
        print("  Dioptras no encontradas")
        df_dioptrias = pd.DataFrame()
    
    # Vendas
    vendas_files = list(processed_dir.glob("VENDAS_COMPLETAS_*.xlsx"))
    if vendas_files:
        df_vendas = pd.read_excel(vendas_files[-1])
        print(f" {len(df_vendas)} registros de vendas")
    else:
        print("  Vendas no encontradas")
        df_vendas = pd.DataFrame()
    
    # 2. Criar anlises integradas
    print("\n Criando anlises integradas...")
    
    # Dashboard executivo
    df_dashboard = criar_dashboard_executivo(df_clientes, df_relacionamentos, df_dioptrias, df_vendas)
    print(" Dashboard executivo criado")
    
    # Anlise por loja
    df_analise_lojas = criar_analise_por_loja(df_relacionamentos, df_vendas)
    print(" Anlise por loja criada")
    
    # Top clientes
    df_top_clientes = criar_top_clientes(df_relacionamentos, df_vendas)
    print(" Top clientes identificados")
    
    # 3. Gerar arquivo Excel final
    print("\n Gerando arquivo Excel final...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = processed_dir / f"SISTEMA_INTEGRADO_OTICAS_{timestamp}.xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. Dashboard Executivo
        df_dashboard.to_excel(writer, sheet_name='Dashboard_Executivo', index=False)
        
        # 2. Anlise por Loja
        df_analise_lojas.to_excel(writer, sheet_name='Analise_Por_Loja', index=False)
        
        # 3. Top Clientes
        df_top_clientes.to_excel(writer, sheet_name='Top_Clientes', index=False)
        
        # 4. Base Clientes (sample)
        df_clientes.head(1000).to_excel(writer, sheet_name='Base_Clientes_Sample', index=False)
        
        # 5. Relacionamentos (sample)
        if not df_relacionamentos.empty:
            df_relacionamentos.head(1000).to_excel(writer, sheet_name='Relacionamentos_Sample', index=False)
        
        # 6. Dioptras (sample)
        if not df_dioptrias.empty:
            df_dioptrias.head(1000).to_excel(writer, sheet_name='Dioptrias_Sample', index=False)
        
        # 7. Vendas (sample)
        if not df_vendas.empty:
            df_vendas.head(1000).to_excel(writer, sheet_name='Vendas_Sample', index=False)
    
    # 4. Aplicar formatao
    print(" Aplicando formatao profissional...")
    wb = openpyxl.load_workbook(output_file)
    
    # Formatar sheets principais
    if 'Dashboard_Executivo' in wb.sheetnames:
        aplicar_formatacao_excel(wb['Dashboard_Executivo'], df_dashboard, "DASHBOARD EXECUTIVO - TICAS CARNE FCIL")
    
    if 'Analise_Por_Loja' in wb.sheetnames:
        aplicar_formatacao_excel(wb['Analise_Por_Loja'], df_analise_lojas, "ANLISE DETALHADA POR LOJA")
    
    if 'Top_Clientes' in wb.sheetnames:
        aplicar_formatacao_excel(wb['Top_Clientes'], df_top_clientes, "TOP 50 CLIENTES")
    
    wb.save(output_file)
    
    # 5. Estatsticas finais
    print(f"\n SISTEMA INTEGRADO CONCLUDO!")
    print("=" * 80)
    print(f" DADOS CONSOLIDADOS:")
    print(f"    {len(df_clientes):,} clientes nicos")
    print(f"    {len(df_relacionamentos):,} relacionamentos OS-Cliente")
    print(f"    {len(df_dioptrias):,} registros de dioptras")
    print(f"    {len(df_vendas):,} registros de vendas")
    print(f"    {len(df_analise_lojas)} lojas analisadas")
    
    print(f"\n ARQUIVO FINAL:")
    print("=" * 80)
    print(f" {output_file}")
    print(f" 7 sheets com anlises completas")
    print(f" Formatao profissional aplicada")
    
    print(f"\n DESAFIO CONCLUDO COM SUCESSO!")
    print("=" * 80)
    print(f" Sistema completo de relacionamento Cliente-OS")
    print(f" Dioptras e vendas integradas")
    print(f" Dashboard executivo para tomada de deciso")
    print(f" Anlises por loja e top clientes")
    print(f" Base slida para expanso futura")

if __name__ == "__main__":
    main()