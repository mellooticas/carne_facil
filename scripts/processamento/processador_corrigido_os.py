#!/usr/bin/env python3
"""
PROCESSADOR CORRIGIDO - AGRUPAMENTO POR OS ÚNICA
Versão corrigida que agrupa corretamente OS com múltiplas formas de pagamento
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import re

class ProcessadorCorrigidoOS:
    def __init__(self):
        self.pasta_dados = Path("data/caixa_lojas")
        self.pasta_corrigidos = Path("data/documentos_corrigidos")
        self.pasta_corrigidos.mkdir(parents=True, exist_ok=True)
        
        self.colunas_padrao = ['loja', 'data', 'numero_os', 'cliente', 'formas_pagamento', 'valor_total', 'entrada']
        self.lojas_ativas = ['MAUA', 'SUZANO', 'RIO_PEQUENO', 'PERUS', 'SUZANO2', 'SAO_MATEUS']
    
    def extrair_vendas_arquivo_corrigido(self, arquivo_path, loja):
        """Extrai vendas agrupando corretamente por número de OS"""
        print(f"   📄 Processando: {arquivo_path.name}")
        
        try:
            wb = openpyxl.load_workbook(arquivo_path, data_only=True)
            vendas_dia = {}  # Dicionário para agrupar por OS
            
            # Detectar ano e mês
            nome_arquivo = arquivo_path.name.lower()
            meses = {'jan': 'JAN', 'fev': 'FEV', 'mar': 'MAR', 'abr': 'ABR', 
                    'mai': 'MAI', 'jun': 'JUN', 'jul': 'JUL', 'ago': 'AGO',
                    'set': 'SET', 'out': 'OUT', 'nov': 'NOV', 'dez': 'DEZ'}
            
            mes_nome = 'UNKNOWN'
            for mes_abr, mes_nome_full in meses.items():
                if mes_abr in nome_arquivo:
                    mes_nome = mes_nome_full
                    break
            
            if '_25' in nome_arquivo or '2025' in nome_arquivo:
                ano = '2025'
            elif '_24' in nome_arquivo or '2024' in nome_arquivo:
                ano = '2024'
            elif '_23' in nome_arquivo or '2023' in nome_arquivo:
                ano = '2023'
            else:
                ano = '2024'
            
            # Processar dias
            for dia in range(1, 32):
                dia_str = f"{dia:02d}"
                
                if dia_str in wb.sheetnames:
                    ws = wb[dia_str]
                    
                    # Buscar dados na região E-G
                    for row in range(5, ws.max_row + 1):
                        numero_os = ws.cell(row=row, column=5).value  # Coluna E
                        cliente = ws.cell(row=row, column=6).value    # Coluna F
                        forma_pgto = ws.cell(row=row, column=7).value # Coluna G
                        valor = ws.cell(row=row, column=8).value      # Coluna H (se existir)
                        
                        if numero_os and str(numero_os).strip():
                            numero_os = str(numero_os).strip()
                            cliente_nome = str(cliente).strip() if cliente else 'CLIENTE_NAO_INFORMADO'
                            forma_pgto_str = str(forma_pgto).strip() if forma_pgto else 'NAO_INFORMADO'
                            valor_num = float(valor) if valor and str(valor).replace('.', '').replace(',', '').isdigit() else 0
                            
                            # Chave única por OS
                            chave_os = f"{loja}_{numero_os}_{dia:02d}_{mes_nome}_{ano}"
                            
                            if chave_os in vendas_dia:
                                # OS já existe - agregar formas de pagamento
                                os_existente = vendas_dia[chave_os]
                                
                                # Concatenar formas de pagamento
                                formas_existentes = os_existente['formas_pagamento'].split('+')
                                if forma_pgto_str not in formas_existentes and forma_pgto_str != 'NAO_INFORMADO':
                                    os_existente['formas_pagamento'] += f"+{forma_pgto_str}"
                                
                                # Somar valor se maior que zero
                                if valor_num > 0:
                                    if os_existente['valor_total'] == 0:
                                        os_existente['valor_total'] = valor_num
                                    else:
                                        # Verificar se devemos somar ou manter o maior
                                        # Para casos como DN+CTC, geralmente o primeiro valor é o total
                                        pass  # Manter valor original
                                
                            else:
                                # Nova OS
                                vendas_dia[chave_os] = {
                                    'loja': loja,
                                    'data': f"{dia:02d}/{mes_nome.lower()[:3]}/{ano}",
                                    'numero_os': numero_os,
                                    'cliente': cliente_nome,
                                    'formas_pagamento': forma_pgto_str,
                                    'valor_total': valor_num,
                                    'entrada': 0  # Será calculado depois se necessário
                                }
            
            wb.close()
            
            # Converter para lista
            vendas_consolidadas = list(vendas_dia.values())
            
            print(f"      ✅ {len(vendas_consolidadas)} OS únicas extraídas")
            return vendas_consolidadas
            
        except Exception as e:
            print(f"      ❌ Erro: {e}")
            return []
    
    def processar_loja_corrigida(self, loja):
        """Processa loja com agrupamento correto de OS"""
        print(f"\n🔧 PROCESSANDO LOJA CORRIGIDA: {loja}")
        print("=" * 60)
        
        # Detectar arquivos da loja
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            print(f"❌ Loja {loja} não encontrada")
            return None
        
        todas_vendas = []
        arquivos_processados = 0
        
        # Processar arquivos da raiz (2025)
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~'):
                vendas = self.extrair_vendas_arquivo_corrigido(arquivo, loja)
                todas_vendas.extend(vendas)
                if vendas:
                    arquivos_processados += 1
        
        # Processar pastas de anos
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir() and any(ano in pasta.name for ano in ['2023', '2024', '2025']):
                for arquivo in pasta.glob("*.xlsx"):
                    if not arquivo.name.startswith('~'):
                        vendas = self.extrair_vendas_arquivo_corrigido(arquivo, loja)
                        todas_vendas.extend(vendas)
                        if vendas:
                            arquivos_processados += 1
        
        if todas_vendas:
            # Criar DataFrame
            df = pd.DataFrame(todas_vendas)
            
            # Remover duplicatas baseado em chave única
            df['chave_unica'] = df['loja'] + '_' + df['numero_os'] + '_' + df['data']
            antes = len(df)
            df = df.drop_duplicates(subset=['chave_unica'])
            depois = len(df)
            
            if antes != depois:
                print(f"   🧹 Removidas {antes - depois} OS duplicadas")
            
            # Salvar documento corrigido
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"OS_CORRIGIDAS_{loja}_{timestamp}.xlsx"
            caminho_arquivo = self.pasta_corrigidos / nome_arquivo
            
            df.to_excel(caminho_arquivo, index=False)
            
            print(f"\n📊 RESULTADO CORRIGIDO:")
            print(f"   ✅ Arquivos processados: {arquivos_processados}")
            print(f"   📈 OS únicas: {len(df):,}")
            print(f"   💰 Valor total: R$ {df['valor_total'].sum():,.2f}")
            print(f"   📅 Período: {df['data'].min()} a {df['data'].max()}")
            print(f"   💾 Arquivo salvo: {nome_arquivo}")
            
            return {
                'loja': loja,
                'arquivo': nome_arquivo,
                'os_unicas': len(df),
                'valor_total': df['valor_total'].sum(),
                'diferenca_anterior': antes - depois
            }
        
        return None
    
    def processar_todas_lojas_corrigidas(self):
        """Processa todas as lojas com correção"""
        print("🔧 PROCESSAMENTO CORRIGIDO - TODAS AS LOJAS")
        print("=" * 70)
        print("🎯 Agrupamento correto por número de OS única")
        print()
        
        resultados = []
        
        for loja in self.lojas_ativas:
            resultado = self.processar_loja_corrigida(loja)
            if resultado:
                resultados.append(resultado)
        
        if resultados:
            print(f"\n🎉 PROCESSAMENTO CORRIGIDO CONCLUÍDO!")
            print("=" * 50)
            
            total_os = sum(r['os_unicas'] for r in resultados)
            total_valor = sum(r['valor_total'] for r in resultados)
            total_correcoes = sum(r['diferenca_anterior'] for r in resultados)
            
            print(f"📊 COMPARAÇÃO DOS RESULTADOS:")
            print(f"   🏢 Lojas processadas: {len(resultados)}")
            print(f"   📈 OS únicas (corrigido): {total_os:,}")
            print(f"   💰 Valor total: R$ {total_valor:,.2f}")
            print(f"   🔧 Total de correções: {total_correcoes}")
            
            print(f"\n📄 ARQUIVOS CORRIGIDOS GERADOS:")
            for resultado in resultados:
                print(f"   ✅ {resultado['loja']}: {resultado['arquivo']}")
                print(f"      📊 {resultado['os_unicas']} OS | Correções: {resultado['diferenca_anterior']}")
        
        return resultados

def main():
    processador = ProcessadorCorrigidoOS()
    
    print("🔧 PROCESSADOR CORRIGIDO - AGRUPAMENTO POR OS ÚNICA")
    print("=" * 65)
    print("🎯 Corrige problema de OS com múltiplas formas de pagamento")
    print()
    print("1. Processar todas as lojas (versão corrigida)")
    print("2. Processar loja específica")
    print("3. Sair")
    
    while True:
        escolha = input("\n👉 Escolha uma opção (1-3): ").strip()
        
        if escolha == "1":
            processador.processar_todas_lojas_corrigidas()
            break
        elif escolha == "2":
            loja = input("Digite o nome da loja: ").strip().upper()
            if loja in processador.lojas_ativas:
                processador.processar_loja_corrigida(loja)
            else:
                print(f"❌ Loja {loja} não encontrada")
            break
        elif escolha == "3":
            print("👋 Saindo...")
            break
        else:
            print("❌ Opção inválida")

if __name__ == "__main__":
    main()