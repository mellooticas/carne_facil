#!/usr/bin/env python3
"""
PROCESSADOR REST_ENTR FINAL - Extrair dados estruturados
Processa tabela rest_entr(dia) na posição E24 e gera planilhas finais
Estrutura: Loja | Data | Cliente | Valor_Restante | Informações_Adicionais
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime

class ProcessadorRestEntrFinal:
    def __init__(self):
        self.pasta_dados = Path("data/caixa_lojas")
        self.pasta_resultados = Path("data/rest_entr_final")
        self.pasta_resultados.mkdir(parents=True, exist_ok=True)
        
        self.lojas_ativas = ['MAUA', 'SUZANO', 'RIO_PEQUENO', 'PERUS', 'SUZANO2', 'SAO_MATEUS']
    
    def extrair_rest_entr_estruturado(self, arquivo_path, loja):
        """Extrai dados estruturados da tabela rest_entr na posição E24"""
        print(f"   📄 Processando: {arquivo_path.name}")
        
        try:
            wb = openpyxl.load_workbook(arquivo_path, data_only=True)
            
            # Detectar ano e mês
            nome_arquivo = arquivo_path.name.lower()
            meses = {'jan': 'JAN', 'fev': 'FEV', 'mar': 'MAR', 'abr': 'ABR', 
                    'mai': 'MAI', 'jun': 'JUN', 'jul': 'JUL', 'ago': 'AGO',
                    'set': 'SET', 'out': 'OUT', 'nov': 'NOV', 'dez': 'DEZ'}
            
            mes_nome = 'UNKNOWN'
            for mes_abr, mes_nome_full in meses.items():
                if mes_abr in nome_arquivo:
                    mes_nome = mes_nome_full
                    break
            
            if '_25' in nome_arquivo or '2025' in nome_arquivo:
                ano = '2025'
            elif '_24' in nome_arquivo or '2024' in nome_arquivo:
                ano = '2024'
            elif '_23' in nome_arquivo or '2023' in nome_arquivo:
                ano = '2023'
            else:
                ano = '2024'
            
            dados_rest_entr = []
            
            # Processar cada dia (sheets numeradas)
            for dia in range(1, 32):
                dia_str = f"{dia:02d}"
                
                if dia_str in wb.sheetnames:
                    ws = wb[dia_str]
                    
                    # Verificar se E24 tem "Restante Entrada"
                    cabecalho_e24 = ws.cell(row=24, column=5).value  # E24
                    
                    if cabecalho_e24 and 'restante' in str(cabecalho_e24).lower() and 'entrada' in str(cabecalho_e24).lower():
                        print(f"      📅 Dia {dia}: Cabeçalho encontrado - {cabecalho_e24}")
                        
                        # Extrair dados a partir da linha 25 (depois do cabeçalho)
                        entradas_dia = 0
                        for row in range(25, min(45, ws.max_row + 1)):  # Linhas 25-44
                            # Colunas típicas da tabela rest_entr
                            col_cliente = ws.cell(row=row, column=5).value    # E - Cliente/Informação
                            col_valor = ws.cell(row=row, column=6).value      # F - Valor restante
                            col_obs = ws.cell(row=row, column=7).value        # G - Observações/Data
                            col_extra = ws.cell(row=row, column=8).value      # H - Info adicional
                            
                            # Só processar se tem pelo menos cliente ou valor
                            if (col_cliente and str(col_cliente).strip()) or \
                               (col_valor and str(col_valor).strip()):
                                
                                cliente_nome = str(col_cliente).strip() if col_cliente else 'NAO_INFORMADO'
                                
                                # Tentar converter valor
                                valor_restante = 0
                                if col_valor:
                                    try:
                                        if isinstance(col_valor, (int, float)):
                                            valor_restante = float(col_valor)
                                        else:
                                            valor_str = str(col_valor).replace(',', '.').replace('R$', '').strip()
                                            valor_restante = float(valor_str) if valor_str and valor_str != 'None' else 0
                                    except:
                                        valor_restante = 0
                                
                                observacoes = str(col_obs).strip() if col_obs else ''
                                info_adicional = str(col_extra).strip() if col_extra else ''
                                
                                # Só incluir se tem dados relevantes
                                if cliente_nome != 'NAO_INFORMADO' or valor_restante > 0:
                                    entrada = {
                                        'loja': loja,
                                        'ano': ano,
                                        'mes': mes_nome,
                                        'dia': int(dia),
                                        'data_completa': f"{dia:02d}/{mes_nome.lower()[:3]}/{ano}",
                                        'cliente': cliente_nome,
                                        'valor_restante': valor_restante,
                                        'observacoes': observacoes,
                                        'info_adicional': info_adicional,
                                        'arquivo_origem': arquivo_path.name,
                                        'linha_origem': row
                                    }
                                    
                                    dados_rest_entr.append(entrada)
                                    entradas_dia += 1
                        
                        if entradas_dia > 0:
                            print(f"         ✅ {entradas_dia} entradas restantes extraídas")
            
            wb.close()
            
            print(f"      📊 TOTAL: {len(dados_rest_entr)} entradas restantes")
            print(f"      💰 VALOR: R$ {sum(e['valor_restante'] for e in dados_rest_entr):,.2f}")
            
            return dados_rest_entr
            
        except Exception as e:
            print(f"      ❌ Erro: {e}")
            return []
    
    def processar_loja_rest_entr_final(self, loja):
        """Processa loja completa e gera planilha final"""
        print(f"\n💰 PROCESSANDO REST_ENTR FINAL: {loja}")
        print("=" * 60)
        
        pasta_loja = self.pasta_dados / loja
        
        if not pasta_loja.exists():
            print(f"❌ Loja {loja} não encontrada")
            return None
        
        todas_entradas = []
        arquivos_processados = 0
        
        # Processar arquivos da raiz (2025)
        for arquivo in pasta_loja.glob("*.xlsx"):
            if not arquivo.name.startswith('~'):
                entradas = self.extrair_rest_entr_estruturado(arquivo, loja)
                todas_entradas.extend(entradas)
                if entradas:
                    arquivos_processados += 1
        
        # Processar pastas de anos
        for pasta in pasta_loja.iterdir():
            if pasta.is_dir() and any(ano in pasta.name for ano in ['2023', '2024', '2025']):
                for arquivo in pasta.glob("*.xlsx"):
                    if not arquivo.name.startswith('~'):
                        entradas = self.extrair_rest_entr_estruturado(arquivo, loja)
                        todas_entradas.extend(entradas)
                        if entradas:
                            arquivos_processados += 1
        
        if todas_entradas:
            # Criar DataFrame
            df = pd.DataFrame(todas_entradas)
            
            # Consolidar entradas duplicadas (mesmo cliente, mesmo dia)
            print(f"\n🔧 CONSOLIDANDO ENTRADAS DUPLICADAS...")
            df['chave_unica'] = df['loja'] + '_' + df['cliente'] + '_' + df['data_completa']
            
            antes = len(df)
            df_consolidado = df.groupby('chave_unica').agg({
                'loja': 'first',
                'ano': 'first',
                'mes': 'first',
                'dia': 'first',
                'data_completa': 'first',
                'cliente': 'first',
                'valor_restante': 'sum',  # Somar valores restantes
                'observacoes': lambda x: ' | '.join(x.unique()),
                'info_adicional': lambda x: ' | '.join(x.unique()),
                'arquivo_origem': 'first',
                'linha_origem': 'first'
            }).reset_index()
            
            depois = len(df_consolidado)
            
            if antes != depois:
                print(f"   🧹 Consolidadas {antes - depois} entradas duplicadas")
            
            # Filtrar entradas com valor > 0
            df_com_valor = df_consolidado[df_consolidado['valor_restante'] > 0]
            
            # Salvar resultado
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"REST_ENTR_FINAL_{loja}_{timestamp}.xlsx"
            caminho_arquivo = self.pasta_resultados / nome_arquivo
            
            # Reordenar colunas
            colunas_finais = ['loja', 'ano', 'mes', 'dia', 'data_completa', 'cliente', 
                             'valor_restante', 'observacoes', 'info_adicional', 
                             'arquivo_origem', 'linha_origem']
            df_final = df_consolidado[colunas_finais]
            
            df_final.to_excel(caminho_arquivo, index=False)
            
            print(f"\n📊 RESULTADO REST_ENTR FINAL {loja}:")
            print(f"   ✅ Arquivos processados: {arquivos_processados}")
            print(f"   📈 Entradas totais: {len(df_consolidado):,}")
            print(f"   💰 Valor total restante: R$ {df_consolidado['valor_restante'].sum():,.2f}")
            print(f"   💵 Entradas com valor > 0: {len(df_com_valor):,}")
            print(f"   📅 Período: {df_consolidado['data_completa'].min()} a {df_consolidado['data_completa'].max()}")
            print(f"   💾 Arquivo salvo: {nome_arquivo}")
            
            return {
                'loja': loja,
                'arquivo': nome_arquivo,
                'entradas_totais': len(df_consolidado),
                'entradas_com_valor': len(df_com_valor),
                'valor_total_restante': df_consolidado['valor_restante'].sum(),
                'consolidacoes': antes - depois
            }
        
        return None
    
    def processar_todas_lojas_rest_entr_final(self):
        """Processa todas as lojas e gera planilhas finais"""
        print("💰 PROCESSAMENTO REST_ENTR FINAL - TODAS AS LOJAS")
        print("=" * 70)
        print("🎯 Estrutura: Loja | Data | Cliente | Valor_Restante | Observações")
        print()
        
        resultados = []
        
        for loja in self.lojas_ativas:
            resultado = self.processar_loja_rest_entr_final(loja)
            if resultado:
                resultados.append(resultado)
        
        if resultados:
            print(f"\n🎉 PROCESSAMENTO REST_ENTR FINAL CONCLUÍDO!")
            print("=" * 50)
            
            total_entradas = sum(r['entradas_totais'] for r in resultados)
            total_com_valor = sum(r['entradas_com_valor'] for r in resultados)
            total_valor_restante = sum(r['valor_total_restante'] for r in resultados)
            total_consolidacoes = sum(r['consolidacoes'] for r in resultados)
            
            print(f"📊 RESULTADO FINAL REST_ENTR:")
            print(f"   🏢 Lojas processadas: {len(resultados)}")
            print(f"   📈 Entradas totais: {total_entradas:,}")
            print(f"   💵 Entradas com valor: {total_com_valor:,}")
            print(f"   💰 Valor total restante: R$ {total_valor_restante:,.2f}")
            print(f"   🔧 Consolidações aplicadas: {total_consolidacoes}")
            
            print(f"\n📄 ARQUIVOS GERADOS:")
            for resultado in resultados:
                print(f"   ✅ {resultado['loja']}: {resultado['arquivo']}")
                print(f"      📊 {resultado['entradas_com_valor']:,} entradas | R$ {resultado['valor_total_restante']:,.2f}")
        
        return resultados

def main():
    processador = ProcessadorRestEntrFinal()
    
    print("💰 PROCESSADOR REST_ENTR FINAL")
    print("=" * 50)
    print("🎯 Gera planilhas finais com estrutura: Loja | Data | Cliente | Valor")
    print("📍 Extrai dados da posição E24 (Restante Entrada)")
    print()
    print("1. Processar todas as lojas (planilhas finais)")
    print("2. Processar loja específica")
    print("3. Sair")
    
    while True:
        escolha = input("\n👉 Escolha uma opção (1-3): ").strip()
        
        if escolha == "1":
            processador.processar_todas_lojas_rest_entr_final()
            break
        elif escolha == "2":
            loja = input("Digite o nome da loja: ").strip().upper()
            if loja in processador.lojas_ativas:
                processador.processar_loja_rest_entr_final(loja)
            else:
                print(f"❌ Loja {loja} não encontrada")
            break
        elif escolha == "3":
            print("👋 Saindo...")
            break
        else:
            print("❌ Opção inválida")

if __name__ == "__main__":
    main()