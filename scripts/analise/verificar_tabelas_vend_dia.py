#!/usr/bin/env python3
"""
VERIFICADOR DE TABELAS EXCEL - FOCO NA VEND(DIA)
Análise da estrutura das tabelas Excel para processar apenas vend(dia)
"""

import pandas as pd
import openpyxl
from pathlib import Path

def verificar_estrutura_excel(arquivo_path):
    """Verifica estrutura completa de um arquivo Excel"""
    print(f"\n📄 Analisando: {arquivo_path.name}")
    print("=" * 60)
    
    try:
        # Carregar workbook
        wb = openpyxl.load_workbook(arquivo_path, data_only=True)
        
        print(f"📋 PLANILHAS DISPONÍVEIS ({len(wb.sheetnames)}):")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i:2d}. {sheet_name}")
        
        # Verificar se existe vend(dia)
        if 'vend(dia)' in wb.sheetnames:
            print(f"\n✅ TABELA 'vend(dia)' ENCONTRADA!")
            
            ws = wb['vend(dia)']
            
            # Verificar estrutura da tabela
            print(f"\n📊 ESTRUTURA DA TABELA vend(dia):")
            print(f"   📐 Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
            
            # Verificar cabeçalhos (primeiras 3 linhas)
            print(f"\n📝 CABEÇALHOS/ESTRUTURA:")
            for row in range(1, min(4, ws.max_row + 1)):
                linha_dados = []
                for col in range(1, min(10, ws.max_column + 1)):  # Primeiras 10 colunas
                    valor = ws.cell(row=row, column=col).value
                    if valor:
                        linha_dados.append(f"[{col}]{valor}")
                
                if linha_dados:
                    print(f"   Linha {row}: {' | '.join(linha_dados)}")
            
            # Verificar dados de exemplo (linhas 4-8)
            print(f"\n🔍 DADOS DE EXEMPLO:")
            for row in range(4, min(9, ws.max_row + 1)):
                linha_dados = []
                for col in range(1, min(8, ws.max_column + 1)):
                    valor = ws.cell(row=row, column=col).value
                    if valor:
                        linha_dados.append(f"[{col}]{valor}")
                
                if linha_dados:
                    print(f"   Linha {row}: {' | '.join(linha_dados)}")
            
            # Buscar padrões de OS/Cliente/Valor
            print(f"\n🎯 ANÁLISE DE PADRÕES:")
            os_encontradas = []
            valores_encontrados = []
            
            for row in range(1, min(20, ws.max_row + 1)):
                for col in range(1, ws.max_column + 1):
                    valor = ws.cell(row=row, column=col).value
                    
                    # Buscar padrões de OS (números)
                    if isinstance(valor, (int, float)) and valor > 1000:
                        os_encontradas.append(f"Linha {row}, Col {col}: {valor}")
                    
                    # Buscar valores monetários
                    if isinstance(valor, (int, float)) and 10 <= valor <= 5000:
                        valores_encontrados.append(f"Linha {row}, Col {col}: R$ {valor}")
            
            if os_encontradas:
                print(f"   📈 Possíveis OS encontradas:")
                for os in os_encontradas[:5]:  # Primeiras 5
                    print(f"      {os}")
            
            if valores_encontrados:
                print(f"   💰 Possíveis valores encontrados:")
                for valor in valores_encontrados[:5]:  # Primeiros 5
                    print(f"      {valor}")
        
        else:
            print(f"\n❌ TABELA 'vend(dia)' NÃO ENCONTRADA!")
            print(f"   🔍 Verificando tabelas alternativas...")
            
            # Buscar tabelas similares
            tabelas_similares = []
            for sheet in wb.sheetnames:
                if any(palavra in sheet.lower() for palavra in ['vend', 'dia', 'vendas', 'caixa']):
                    tabelas_similares.append(sheet)
            
            if tabelas_similares:
                print(f"   📋 Tabelas similares encontradas:")
                for tabela in tabelas_similares:
                    print(f"      ✓ {tabela}")
        
        wb.close()
        return 'vend(dia)' in wb.sheetnames
        
    except Exception as e:
        print(f"   ❌ ERRO: {e}")
        return False

def verificar_todas_lojas():
    """Verifica estrutura de todas as lojas"""
    pasta_dados = Path("data/caixa_lojas")
    lojas_com_vend_dia = []
    total_arquivos = 0
    
    print("🔍 VERIFICAÇÃO COMPLETA - FOCO NA TABELA vend(dia)")
    print("=" * 70)
    
    for loja_pasta in pasta_dados.iterdir():
        if loja_pasta.is_dir():
            loja_nome = loja_pasta.name
            print(f"\n🏢 LOJA: {loja_nome}")
            print("-" * 40)
            
            arquivos_com_vend = 0
            arquivos_processados = 0
            
            # Verificar arquivos na raiz
            for arquivo in loja_pasta.glob("*.xlsx"):
                if not arquivo.name.startswith('~'):
                    total_arquivos += 1
                    arquivos_processados += 1
                    tem_vend_dia = verificar_estrutura_excel(arquivo)
                    if tem_vend_dia:
                        arquivos_com_vend += 1
            
            # Verificar pastas de anos
            for pasta_ano in loja_pasta.iterdir():
                if pasta_ano.is_dir():
                    for arquivo in pasta_ano.glob("*.xlsx"):
                        if not arquivo.name.startswith('~'):
                            total_arquivos += 1
                            arquivos_processados += 1
                            tem_vend_dia = verificar_estrutura_excel(arquivo)
                            if tem_vend_dia:
                                arquivos_com_vend += 1
            
            if arquivos_com_vend > 0:
                lojas_com_vend_dia.append({
                    'loja': loja_nome,
                    'arquivos_com_vend': arquivos_com_vend,
                    'total_arquivos': arquivos_processados
                })
            
            print(f"\n📊 RESUMO {loja_nome}:")
            print(f"   ✅ Arquivos com vend(dia): {arquivos_com_vend}")
            print(f"   📁 Total de arquivos: {arquivos_processados}")
    
    print(f"\n🎉 RESULTADO FINAL:")
    print("=" * 50)
    print(f"📁 Total de arquivos verificados: {total_arquivos}")
    print(f"🏢 Lojas com tabela vend(dia):")
    
    for loja_info in lojas_com_vend_dia:
        print(f"   ✅ {loja_info['loja']}: {loja_info['arquivos_com_vend']}/{loja_info['total_arquivos']} arquivos")
    
    return lojas_com_vend_dia

if __name__ == "__main__":
    # Primeiro teste com um arquivo específico
    print("🧪 TESTE INICIAL - VERIFICANDO UM ARQUIVO")
    print("=" * 50)
    
    arquivo_teste = Path("data/caixa_lojas/SUZANO/2024_SUZ/jan_24.xlsx")
    if arquivo_teste.exists():
        verificar_estrutura_excel(arquivo_teste)
    else:
        print("❌ Arquivo de teste não encontrado")
    
    print("\n" + "="*70)
    input("Pressione ENTER para verificar todas as lojas...")
    
    # Verificação completa
    lojas_resultado = verificar_todas_lojas()
    
    if lojas_resultado:
        print(f"\n✅ Sistema pronto para processar {len(lojas_resultado)} lojas com tabela vend(dia)!")
    else:
        print(f"\n❌ Nenhuma loja com tabela vend(dia) encontrada!")